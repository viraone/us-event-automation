"""Read-only spreadsheet data loader for the local US Event dashboard.

The scraper remains the source of truth.  This module discovers every matching
``US_Event_*.xlsx`` workbook in the project directory and builds a clean,
JSON-friendly snapshot for Flask/Jinja without modifying the workbooks.

The public entry point is :func:`load_dashboard_data`.  It deliberately uses
only ``openpyxl`` and the Python standard library so the dashboard can refresh
from disk on every request without a database or a separate import step.
"""

from __future__ import annotations

from collections import Counter
from datetime import date, datetime, time, timedelta
from pathlib import Path
import calendar
import hashlib
import math
import re
from typing import Any, Mapping
from urllib.parse import quote

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


WORKBOOK_PATTERN = "US_Event_*.xlsx"
PREFERRED_CITY_PRIORITIES = {
    "bellevue": 1,
    "kirkland": 2,
    "seattle": 3,
}
VALID_STATUSES = {"NEW", "EXISTING", "REMOVED"}

_MONTH_NAME_TO_NUMBER = {
    name.casefold(): number
    for number, name in enumerate(calendar.month_name)
    if name
}
_FILE_MONTH_RE = re.compile(
    r"^US_Event_(?P<month>[A-Za-z]+)_(?P<year>\d{4})(?:_Gigs)?\.xlsx$",
    re.IGNORECASE,
)
_TIME_RE = re.compile(
    r"^\s*(?P<hour>\d{1,2})(?::(?P<minute>\d{1,2}))?\s*"
    r"(?P<period>a|p|am|pm|a\.m\.|p\.m\.)?\s*$",
    re.IGNORECASE,
)
_RAW_TIME_RE = re.compile(
    r"^\s*(?P<hour>\d{1,2})(?::(?P<minute>\d{1,2}))?\s*"
    r"(?P<period>a|p|am|pm|a\.m\.|p\.m\.)\b",
    re.IGNORECASE,
)


def _header_token(value: Any) -> str:
    """Return a punctuation-insensitive token used to match column names."""

    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().casefold())


_HEADER_ALIASES = {
    "date": {"date", "gigdate", "eventdate"},
    "day": {"day", "weekday", "dayofweek"},
    "start_time": {"starttime", "time", "eventtime", "gigtime"},
    "start_time_sort": {"starttimesort", "timesort", "sorttime"},
    "state": {"state", "province", "region"},
    "store_account": {
        "storeaccount",
        "accountstore",
        "account",
        "store",
        "storeaccountname",
    },
    "store_number": {"storenumber", "storeno", "storeid", "locationnumber"},
    "city": {"city", "locationcity", "town"},
    "preferred_area": {"preferredarea", "preferred", "preferredcity"},
    "location_priority": {"locationpriority", "citypriority", "priority"},
    "raw_calendar_listing": {
        "rawcalendarlisting",
        "rawlisting",
        "calendarlisting",
        "eventlisting",
        "rawtext",
    },
    "calendar_color": {"calendarcolor", "eventcolor", "color"},
    "event_url": {"eventurl", "calendarurl", "url", "link"},
    "status": {"status", "gigstatus", "eventstatus"},
    "first_seen": {"firstseen", "firstdiscovered", "dateadded"},
    "last_seen": {"lastseen", "lastobserved"},
    "scraped_at": {"scrapedat", "scrapetime", "refreshedat", "updatedat"},
    "event_id": {"eventid", "calendarid", "gigidentifier", "eventidentifier"},
}
_HEADER_LOOKUP = {
    alias: canonical
    for canonical, aliases in _HEADER_ALIASES.items()
    for alias in aliases
}


def _clean_text(value: Any) -> str:
    """Normalize a spreadsheet value into compact display text."""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "YES" if value else "NO"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def _json_value(value: Any) -> Any:
    """Convert optional/unknown workbook values to JSON-safe primitives."""

    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return value.total_seconds()
    return str(value)


def _from_excel_safely(value: Any, epoch: datetime) -> Any:
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return None
    if not math.isfinite(float(value)):
        return None
    try:
        return from_excel(value, epoch)
    except (TypeError, ValueError, OverflowError):
        return None


def _parse_date(value: Any, epoch: datetime) -> date | None:
    """Parse Excel, ISO, and common US date representations."""

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    excel_value = _from_excel_safely(value, epoch)
    if isinstance(excel_value, datetime):
        return excel_value.date()
    if isinstance(excel_value, date):
        return excel_value

    text = _clean_text(value)
    if not text:
        return None
    iso_candidate = text[:-1] + "+00:00" if text.endswith("Z") else text
    try:
        return datetime.fromisoformat(iso_candidate).date()
    except ValueError:
        pass
    for pattern in (
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y/%m/%d",
        "%B %d, %Y",
        "%b %d, %Y",
        "%B %d %Y",
        "%b %d %Y",
    ):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _parse_clock(value: Any, epoch: datetime) -> time | None:
    """Parse Excel time values plus forms such as ``4p`` and ``4:30 PM``."""

    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None, second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(tzinfo=None, second=0, microsecond=0)
    if isinstance(value, timedelta):
        minutes = round(value.total_seconds() / 60) % (24 * 60)
        return time(minutes // 60, minutes % 60)

    # Excel stores ordinary times as a fraction of one day.  A whole number
    # from an evolving schema is more plausibly an hour (for example ``16``)
    # than an Excel date whose time portion happens to be midnight.
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if 1 <= number <= 23:
            hour = int(number)
            minute = round((number - hour) * 60)
            if minute == 60:
                hour = (hour + 1) % 24
                minute = 0
            return time(hour, minute)

    excel_value = _from_excel_safely(value, epoch)
    if isinstance(excel_value, datetime):
        return excel_value.time().replace(second=0, microsecond=0)
    if isinstance(excel_value, time):
        return excel_value.replace(second=0, microsecond=0)
    text = _clean_text(value).casefold()
    if not text:
        return None
    text = text.replace("noon", "12pm").replace("midnight", "12am")
    match = _TIME_RE.match(text)
    if not match:
        return None
    hour = int(match.group("hour"))
    minute = int(match.group("minute") or 0)
    period = (match.group("period") or "").replace(".", "").casefold()
    if minute > 59:
        return None
    if period:
        if not 1 <= hour <= 12:
            return None
        if period.startswith("p") and hour != 12:
            hour += 12
        elif period.startswith("a") and hour == 12:
            hour = 0
    elif hour > 23:
        return None
    return time(hour, minute)


def _parse_raw_clock(raw_listing: Any) -> time | None:
    """Read a leading compact time from raw calendar text as a last resort."""

    match = _RAW_TIME_RE.match(_clean_text(raw_listing))
    if not match:
        return None
    period = match.group("period")
    return _parse_clock(
        f"{match.group('hour')}:{match.group('minute') or '00'}{period}",
        datetime(1899, 12, 30),
    )


def _parse_timestamp(value: Any, epoch: datetime) -> datetime | None:
    """Parse a timestamp without inventing a timezone absent from the sheet."""

    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    excel_value = _from_excel_safely(value, epoch)
    if isinstance(excel_value, datetime):
        return excel_value.replace(tzinfo=None)
    if isinstance(excel_value, date):
        return datetime.combine(excel_value, time.min)

    text = _clean_text(value)
    if not text:
        return None
    iso_candidate = text[:-1] + "+00:00" if text.endswith("Z") else text
    try:
        parsed = datetime.fromisoformat(iso_candidate)
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone().replace(tzinfo=None)
        return parsed
    except ValueError:
        pass
    for pattern in (
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d %I:%M %p",
        "%b %d, %Y %I:%M %p",
        "%B %d, %Y %I:%M %p",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    return None


def _display_time(value: time | None) -> str:
    if value is None:
        return ""
    hour = value.hour % 12 or 12
    period = "AM" if value.hour < 12 else "PM"
    return f"{hour}:{value.minute:02d} {period}"


def _display_date(value: date | None) -> str:
    if value is None:
        return ""
    return f"{calendar.month_abbr[value.month]} {value.day}, {value.year}"


def _display_timestamp(value: datetime | None) -> str:
    if value is None:
        return ""
    hour = value.hour % 12 or 12
    period = "AM" if value.hour < 12 else "PM"
    return (
        f"{calendar.month_abbr[value.month]} {value.day}, {value.year}, "
        f"{hour}:{value.minute:02d} {period}"
    )


def _month_key(value: date | None) -> str:
    return value.strftime("%Y-%m") if value else ""


def _month_label(key: str) -> str:
    try:
        year_text, month_text = key.split("-", 1)
        month_number = int(month_text)
        return f"{calendar.month_name[month_number]} {int(year_text)}"
    except (ValueError, IndexError):
        return key


def _month_from_filename(filename: str) -> tuple[str, str]:
    match = _FILE_MONTH_RE.match(filename)
    if not match:
        return "", ""
    month_number = _MONTH_NAME_TO_NUMBER.get(match.group("month").casefold())
    if not month_number:
        return "", ""
    key = f"{int(match.group('year')):04d}-{month_number:02d}"
    return key, _month_label(key)


def _normalize_city(value: Any) -> str:
    """Normalize whitespace/case while preserving intentional punctuation."""

    text = _clean_text(value)
    if not text:
        return ""
    # Spreadsheet city values are usually title case; repair all-upper/lower
    # values while leaving mixed-case names such as McMinnville untouched.
    if text.isupper() or text.islower():
        return text.title()
    return text


def _normalize_status(value: Any) -> str:
    text = _clean_text(value).upper()
    aliases = {
        "ACTIVE": "EXISTING",
        "CURRENT": "EXISTING",
        "UNCHANGED": "EXISTING",
        "ADDED": "NEW",
        "NEW GIG": "NEW",
        "DELETED": "REMOVED",
        "MISSING": "REMOVED",
    }
    return aliases.get(text, text)


def _preferred_details(city: str, sheet_value: Any) -> tuple[str, int]:
    city_key = city.casefold()
    if city_key in PREFERRED_CITY_PRIORITIES:
        return "YES", PREFERRED_CITY_PRIORITIES[city_key]
    if not city:
        preferred_text = _clean_text(sheet_value).upper()
        if preferred_text in {"YES", "Y", "TRUE", "1"}:
            return "YES", 9
    return "NO", 9


def _format_size(byte_count: int) -> str:
    if byte_count < 1024:
        return f"{byte_count} B"
    if byte_count < 1024 * 1024:
        return f"{byte_count / 1024:.1f} KB"
    return f"{byte_count / (1024 * 1024):.1f} MB"


def _find_header_row(
    worksheet: Any, *, scan_rows: int = 25
) -> tuple[int, list[Any], dict[int, str], dict[int, str]]:
    """Locate an evolving schema rather than assuming headers are on row 1."""

    best: tuple[int, int, list[Any], dict[int, str], dict[int, str]] | None = None
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1
    ):
        values = list(row)
        canonical: dict[int, str] = {}
        original: dict[int, str] = {}
        for index, value in enumerate(values):
            text = _clean_text(value)
            if not text:
                continue
            original[index] = text
            matched = _HEADER_LOOKUP.get(_header_token(text))
            if matched and matched not in canonical.values():
                canonical[index] = matched
        score = len(canonical)
        # A real gig header should expose several semantic columns.  Requiring
        # two avoids mistaking a title row containing only "Date" for headers.
        if score >= 2 and (best is None or score > best[0]):
            best = (score, row_number, values, canonical, original)
    if best is None:
        raise ValueError("Could not locate a recognizable gig header row")
    _, row_number, values, canonical, original = best
    return row_number, values, canonical, original


def _row_has_content(values: Mapping[str, Any]) -> bool:
    return any(
        _clean_text(values.get(name))
        for name in (
            "date",
            "start_time",
            "store_account",
            "city",
            "raw_calendar_listing",
            "event_id",
            "status",
        )
    )


def _build_store_display(account: str, number: str) -> str:
    if not account:
        return number
    if not number:
        return account
    # Do not repeat a store number already embedded in the account name.
    if re.search(rf"(?:^|\D){re.escape(number)}(?:\D|$)", account):
        return account
    return f"{account} {number}"


def _normalize_row(
    raw: Mapping[str, Any],
    extras: Mapping[str, Any],
    *,
    epoch: datetime,
    source_file: str,
    source_row: int,
    source_modified: datetime,
    file_month_key: str,
) -> dict[str, Any]:
    gig_date = _parse_date(raw.get("date"), epoch)
    raw_listing = _clean_text(raw.get("raw_calendar_listing"))
    start_clock = (
        _parse_clock(raw.get("start_time"), epoch)
        or _parse_clock(raw.get("start_time_sort"), epoch)
        or _parse_raw_clock(raw_listing)
    )
    first_seen_dt = _parse_timestamp(raw.get("first_seen"), epoch)
    last_seen_dt = _parse_timestamp(raw.get("last_seen"), epoch)
    scraped_at_dt = _parse_timestamp(raw.get("scraped_at"), epoch)
    city = _normalize_city(raw.get("city"))
    preferred_area, location_priority = _preferred_details(
        city, raw.get("preferred_area")
    )
    status = _normalize_status(raw.get("status"))
    state = _clean_text(raw.get("state")).upper()
    account = _clean_text(raw.get("store_account"))
    store_number = _clean_text(raw.get("store_number"))
    event_id = _clean_text(raw.get("event_id"))
    row_month_key = _month_key(gig_date) or file_month_key
    day = _clean_text(raw.get("day")) or (
        gig_date.strftime("%A") if gig_date else ""
    )
    time_sort = start_clock.strftime("%H:%M") if start_clock else ""
    timestamp_candidates = [
        item for item in (scraped_at_dt, last_seen_dt, first_seen_dt) if item
    ]
    freshness = max(timestamp_candidates, default=source_modified)

    return {
        "date_iso": gig_date.isoformat() if gig_date else "",
        "date_display": _display_date(gig_date),
        "date_full": (
            f"{day}, {calendar.month_name[gig_date.month]} "
            f"{gig_date.day}, {gig_date.year}"
            if gig_date
            else ""
        ),
        "day": day,
        "start_time": _display_time(start_clock),
        "start_time_sort": time_sort,
        "state": state,
        "store_account": account,
        "store_number": store_number,
        "store_display": _build_store_display(account, store_number),
        "city": city,
        "preferred_area": preferred_area,
        "location_priority": location_priority,
        "raw_calendar_listing": raw_listing,
        "calendar_color": _clean_text(raw.get("calendar_color")),
        "event_url": _clean_text(raw.get("event_url")),
        "status": status,
        "first_seen": _display_timestamp(first_seen_dt),
        "first_seen_iso": (
            first_seen_dt.isoformat(timespec="seconds") if first_seen_dt else ""
        ),
        "last_seen": _display_timestamp(last_seen_dt),
        "last_seen_iso": (
            last_seen_dt.isoformat(timespec="seconds") if last_seen_dt else ""
        ),
        "scraped_at": _display_timestamp(scraped_at_dt),
        "scraped_at_iso": (
            scraped_at_dt.isoformat(timespec="seconds") if scraped_at_dt else ""
        ),
        "event_id": event_id,
        "month_key": row_month_key,
        "month_label": _month_label(row_month_key),
        "source_file": source_file,
        "source_row": source_row,
        "extra_fields": {name: _json_value(value) for name, value in extras.items()},
        # Private fields are removed before returning the public snapshot.
        "_freshness": freshness,
        "_source_modified": source_modified,
    }


def _dedupe_key(row: Mapping[str, Any]) -> tuple[str, ...]:
    event_id = _clean_text(row.get("event_id")).casefold()
    if event_id:
        return ("event-id", event_id)
    raw_listing = _clean_text(row.get("raw_calendar_listing")).casefold()
    if raw_listing:
        return (
            "composite",
            _clean_text(row.get("date_iso")),
            _clean_text(row.get("start_time_sort")),
            raw_listing,
        )
    # Sparse rows cannot be safely merged without risking two real gigs being
    # collapsed.  Retain them using their stable workbook coordinates.
    return (
        "source-row",
        _clean_text(row.get("source_file")),
        _clean_text(row.get("source_row")),
    )


def _public_gig_key(row: Mapping[str, Any]) -> str:
    """Return a compact, deterministic UI identity without exposing raw data.

    It deliberately mirrors the loader's deduplication identity: a real event
    ID wins, otherwise date + time + raw listing identifies an ordinary gig,
    with workbook coordinates reserved for unusually sparse legacy rows.
    """

    identity = _dedupe_key(row)
    encoded = "\x1f".join(f"{len(part)}:{part}" for part in identity).encode(
        "utf-8"
    )
    return f"gig-{hashlib.sha256(encoded).hexdigest()[:24]}"


def _merge_duplicate(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    """Keep the freshest state while filling blanks from an older duplicate."""

    if incoming["_freshness"] >= existing["_freshness"]:
        primary, secondary = incoming.copy(), existing
    else:
        primary, secondary = existing.copy(), incoming
    for key, value in secondary.items():
        if key == "extra_fields":
            merged_extras = dict(value)
            merged_extras.update(primary.get("extra_fields", {}))
            primary["extra_fields"] = merged_extras
        elif key not in primary or primary[key] in (None, ""):
            primary[key] = value
    sources = set()
    for row in (existing, incoming):
        sources.update(row.get("source_files", []))
        if row.get("source_file"):
            sources.add(str(row["source_file"]))
    primary["source_files"] = sorted(sources)
    return primary


def _row_sort_key(row: Mapping[str, Any]) -> tuple[Any, ...]:
    return (
        row.get("date_iso") or "9999-12-31",
        row.get("start_time_sort") or "99:99",
        int(row.get("location_priority") or 9),
        str(row.get("city") or "").casefold(),
        str(row.get("store_display") or "").casefold(),
        str(row.get("raw_calendar_listing") or "").casefold(),
    )


def discover_workbooks(project_dir: str | Path | None = None) -> list[Path]:
    """Return matching source workbooks in chronological/name order.

    Discovery is intentionally limited to the project directory itself; backup
    subdirectories are not source data and must not create duplicate gigs.
    """

    directory = Path(project_dir or Path(__file__).resolve().parent).resolve()
    paths = [
        path
        for path in directory.glob(WORKBOOK_PATTERN)
        if path.is_file() and not path.name.startswith("~$")
    ]

    def sort_key(path: Path) -> tuple[str, str]:
        key, _ = _month_from_filename(path.name)
        return key or "9999-99", path.name.casefold()

    return sorted(paths, key=sort_key)


def _load_workbook_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Read one workbook in read-only mode and return rows plus file metadata."""

    stat = path.stat()
    source_modified = datetime.fromtimestamp(stat.st_mtime)
    filename_month_key, filename_month_label = _month_from_filename(path.name)
    rows: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        # Prefer the active scraper output sheet but tolerate workbooks whose
        # active sheet changed by selecting the sheet with the strongest header.
        candidate_sheets: list[tuple[int, Any, tuple[Any, ...]]] = []
        for worksheet in workbook.worksheets:
            try:
                header = _find_header_row(worksheet)
            except ValueError:
                continue
            candidate_sheets.append((len(header[2]), worksheet, header))
        if not candidate_sheets:
            raise ValueError("No worksheet contains recognizable gig columns")
        _, worksheet, header_data = max(candidate_sheets, key=lambda item: item[0])
        header_row, header_values, canonical_columns, original_columns = header_data

        for source_row, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            raw = {
                canonical: values[index] if index < len(values) else None
                for index, canonical in canonical_columns.items()
            }
            if not _row_has_content(raw):
                continue
            extras = {
                original_columns[index]: values[index]
                for index in original_columns
                if index not in canonical_columns
                and index < len(values)
                and values[index] is not None
            }
            rows.append(
                _normalize_row(
                    raw,
                    extras,
                    epoch=workbook.epoch,
                    source_file=path.name,
                    source_row=source_row,
                    source_modified=source_modified,
                    file_month_key=filename_month_key,
                )
            )
    finally:
        workbook.close()

    inferred_months = Counter(row["month_key"] for row in rows if row["month_key"])
    month_key = filename_month_key
    if not month_key and inferred_months:
        month_key = inferred_months.most_common(1)[0][0]
    month_label = filename_month_label or _month_label(month_key)
    file_timestamps = [
        row["_freshness"] for row in rows if isinstance(row.get("_freshness"), datetime)
    ]
    latest_data = max(file_timestamps, default=source_modified)
    metadata = {
        "name": path.name,
        "filename": path.name,
        "download_name": path.name,
        "download_url": f"/spreadsheets/{quote(path.name)}",
        "path": str(path),
        "sheet_name": worksheet.title,
        "month_key": month_key,
        "month_label": month_label,
        "row_count": len(rows),
        "size_bytes": stat.st_size,
        "size_display": _format_size(stat.st_size),
        "modified_at": source_modified.isoformat(timespec="seconds"),
        "modified_display": _display_timestamp(source_modified),
        "latest_data_at": latest_data.isoformat(timespec="seconds"),
        "headers": [_clean_text(value) for value in header_values if _clean_text(value)],
        "error": "",
    }
    return rows, metadata


def _build_months(
    rows: list[dict[str, Any]], files: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    keys = {row["month_key"] for row in rows if row["month_key"]}
    keys.update(file["month_key"] for file in files if file.get("month_key"))
    months: list[dict[str, Any]] = []
    for key in sorted(keys):
        month_rows = [row for row in rows if row["month_key"] == key]
        current_rows = [row for row in month_rows if row["status"] != "REMOVED"]
        source_files = sorted(
            {
                file["name"]
                for file in files
                if file.get("month_key") == key and not file.get("error")
            }
        )
        months.append(
            {
                "key": key,
                "value": key,
                "label": _month_label(key),
                "total_count": len(month_rows),
                "current_count": len(current_rows),
                "new_count": sum(row["status"] == "NEW" for row in current_rows),
                "preferred_count": sum(
                    row["preferred_area"] == "YES" for row in current_rows
                ),
                "removed_count": sum(
                    row["status"] == "REMOVED" for row in month_rows
                ),
                "file_names": source_files,
            }
        )
    return months


def _build_filter_options(
    rows: list[dict[str, Any]], months: list[dict[str, Any]]
) -> dict[str, Any]:
    def distinct(field: str) -> list[str]:
        return sorted(
            {str(row[field]) for row in rows if row.get(field)}, key=str.casefold
        )

    start_times = sorted(
        {
            (row["start_time_sort"], row["start_time"])
            for row in rows
            if row.get("start_time")
        }
    )
    return {
        "months": [
            {"value": month["key"], "label": month["label"]} for month in months
        ],
        "dates": sorted({row["date_iso"] for row in rows if row["date_iso"]}),
        "cities": distinct("city"),
        "statuses": [
            status
            for status in ("NEW", "EXISTING", "REMOVED")
            if any(row["status"] == status for row in rows)
        ]
        + sorted(
            {
                row["status"]
                for row in rows
                if row["status"] and row["status"] not in VALID_STATUSES
            }
        ),
        "stores": distinct("store_account"),
        "start_times": [display for _, display in start_times],
        "preferred_areas": [
            value
            for value in ("YES", "NO")
            if any(row["preferred_area"] == value for row in rows)
        ],
        "preferred_cities": ["Bellevue", "Kirkland", "Seattle"],
    }


def load_dashboard_data(
    project_dir: str | Path | None = None, *, strict: bool = False
) -> dict[str, Any]:
    """Build a fresh dashboard snapshot from all matching workbooks.

    Parameters
    ----------
    project_dir:
        Directory containing ``US_Event_*.xlsx``.  By default this is the
        directory containing this module.
    strict:
        When ``False`` (the dashboard default), an unreadable workbook is
        reported in ``errors`` and the remaining files still load.  Set to
        ``True`` in diagnostics/tests to raise the original exception.

    Returns
    -------
    dict
        JSON-friendly keys include ``rows``, ``files``, ``months``, ``summary``,
        ``filter_options``, ``last_refreshed``, and ``errors``.  Rows retain
        optional unknown columns inside ``extra_fields``.
    """

    paths = discover_workbooks(project_dir)
    all_rows: list[dict[str, Any]] = []
    files: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []

    for path in paths:
        try:
            file_rows, metadata = _load_workbook_rows(path)
        except Exception as exc:
            if strict:
                raise
            stat = path.stat()
            source_modified = datetime.fromtimestamp(stat.st_mtime)
            month_key, month_label = _month_from_filename(path.name)
            message = f"{type(exc).__name__}: {exc}"
            metadata = {
                "name": path.name,
                "filename": path.name,
                "download_name": path.name,
                "download_url": f"/spreadsheets/{quote(path.name)}",
                "path": str(path),
                "sheet_name": "",
                "month_key": month_key,
                "month_label": month_label,
                "row_count": 0,
                "size_bytes": stat.st_size,
                "size_display": _format_size(stat.st_size),
                "modified_at": source_modified.isoformat(timespec="seconds"),
                "modified_display": _display_timestamp(source_modified),
                "latest_data_at": source_modified.isoformat(timespec="seconds"),
                "headers": [],
                "error": message,
            }
            errors.append({"file": path.name, "error": message})
            file_rows = []
        files.append(metadata)
        all_rows.extend(file_rows)

    deduplicated: dict[tuple[str, ...], dict[str, Any]] = {}
    for row in all_rows:
        key = _dedupe_key(row)
        if key in deduplicated:
            deduplicated[key] = _merge_duplicate(deduplicated[key], row)
        else:
            new_row = row.copy()
            new_row["source_files"] = [row["source_file"]]
            deduplicated[key] = new_row

    rows = sorted(deduplicated.values(), key=_row_sort_key)
    for row in rows:
        row["gig_key"] = _public_gig_key(row)
        row.pop("_freshness", None)
        row.pop("_source_modified", None)

    months = _build_months(rows, files)
    current_rows = [row for row in rows if row["status"] != "REMOVED"]
    summary = {
        "total_current": len(current_rows),
        "current_gigs": len(current_rows),
        "new": sum(row["status"] == "NEW" for row in current_rows),
        "new_gigs": sum(row["status"] == "NEW" for row in current_rows),
        "preferred": sum(
            row["preferred_area"] == "YES" for row in current_rows
        ),
        "preferred_gigs": sum(
            row["preferred_area"] == "YES" for row in current_rows
        ),
        "removed": sum(row["status"] == "REMOVED" for row in rows),
        "removed_gigs": sum(row["status"] == "REMOVED" for row in rows),
        "total_rows": len(rows),
        "file_count": len(files),
        "month_count": len(months),
    }

    timestamp_values: list[datetime] = []
    for metadata in files:
        for field in ("latest_data_at", "modified_at"):
            value = metadata.get(field)
            if value:
                try:
                    timestamp_values.append(datetime.fromisoformat(str(value)))
                except ValueError:
                    pass
    last_refreshed_dt = max(timestamp_values, default=None)

    return {
        "rows": rows,
        "files": files,
        "months": months,
        "summary": summary,
        "filter_options": _build_filter_options(rows, months),
        "last_refreshed": _display_timestamp(last_refreshed_dt),
        "last_refreshed_iso": (
            last_refreshed_dt.isoformat(timespec="seconds")
            if last_refreshed_dt
            else ""
        ),
        "errors": errors,
        "discovered_file_count": len(paths),
    }


# A descriptive alias for callers that prefer snapshot terminology.
load_dashboard_snapshot = load_dashboard_data


__all__ = [
    "PREFERRED_CITY_PRIORITIES",
    "WORKBOOK_PATTERN",
    "discover_workbooks",
    "load_dashboard_data",
    "load_dashboard_snapshot",
]
