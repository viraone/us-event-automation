"""End-to-end tests for the dashboard's local refresh workflow.

These tests never execute the production ``main.py`` scraper.  Each app is
configured with a disposable Python updater and disposable XLSX files so the
background-process, locking, persistence, rollback, and output-sanitization
contracts can be exercised safely.
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path
import sys
import tempfile
import textwrap
import time
import unittest

from openpyxl import Workbook

from app import create_app


HEADERS = [
    "Date",
    "Day",
    "Start Time",
    "Start Time Sort",
    "State",
    "Store / Account",
    "Store Number",
    "City",
    "Preferred Area",
    "Location Priority",
    "Raw Calendar Listing",
    "Calendar Color",
    "Event URL",
    "Status",
    "First Seen",
    "Last Seen",
    "Scraped At",
    "Event ID",
]


def gig_row(
    *,
    event_id: str,
    status: str = "EXISTING",
    city: str = "Chelan",
    start_time: str = "4p",
    raw_listing: str | None = None,
) -> list[object]:
    """Return one production-shaped workbook row."""

    listing = raw_listing or f"{start_time} WA|| Safeway 3265 ({city})"
    time_sort = {"1p": "13:00", "4p": "16:00"}.get(start_time, "")
    return [
        "2026-08-14",
        "Friday",
        start_time,
        time_sort,
        "WA",
        "Safeway",
        "3265",
        city,
        "YES" if city in {"Bellevue", "Kirkland", "Seattle"} else "NO",
        {"Bellevue": 1, "Kirkland": 2, "Seattle": 3}.get(city, 9),
        listing,
        "#808080",
        "https://example.test/event",
        status,
        "2026-08-10 09:00:00",
        "2026-08-11 09:00:00",
        "2026-08-11 09:00:00",
        event_id,
    ]


def write_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Gigs"
    worksheet.append(HEADERS)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


class RefreshWorkflowTests(unittest.TestCase):
    """Exercise refreshes through Flask exactly as the browser does."""

    def setUp(self) -> None:
        self._temporary_directory = tempfile.TemporaryDirectory()
        self.data_dir = Path(self._temporary_directory.name).resolve()
        self.workbook_path = self.data_dir / "US_Event_August_2026_Gigs.xlsx"
        write_workbook(
            self.workbook_path,
            [gig_row(event_id="existing-1")],
        )

    def tearDown(self) -> None:
        self._temporary_directory.cleanup()

    def make_app(self, update_script: Path):
        app = create_app(data_dir=self.data_dir, update_script=update_script)
        app.config.update(TESTING=True)
        return app

    def write_success_updater(self, *, delay: float = 0.05) -> tuple[Path, Path]:
        """Create a controlled updater that appends exactly one NEW gig."""

        script = self.data_dir / "controlled_success_updater.py"
        marker = self.data_dir / "controlled_success_marker.json"
        script.write_text(
            textwrap.dedent(
                f"""
                import json
                import os
                from pathlib import Path
                import sys
                import time

                from openpyxl import load_workbook

                marker = Path({str(marker)!r})
                workbook_path = Path({str(self.workbook_path)!r})
                marker.write_text(json.dumps({{
                    "argv": sys.argv,
                    "cwd": os.getcwd(),
                    "executable": sys.executable,
                }}), encoding="utf-8")
                print("Connecting to US Event Management...", flush=True)
                time.sleep({delay!r})
                workbook = load_workbook(workbook_path)
                worksheet = workbook.active
                worksheet.append({gig_row(event_id="new-2", status="NEW", city="Seattle", start_time="1p")!r})
                workbook.save(workbook_path)
                workbook.close()
                print("Updating spreadsheets...", flush=True)
                """
            ).lstrip(),
            encoding="utf-8",
        )
        return script, marker

    def write_destructive_failure_updater(self) -> tuple[Path, dict[str, str]]:
        """Create a failing updater that corrupts XLSX bytes before exiting."""

        secrets = {
            "email": "sample.user@example.test",
            "password": "Sup3r-Secret-Password!",
            "token": "tok_live_refresh_workflow_secret",
            "cookie": "sessionid=private-cookie-value",
            "bearer": "bearer-auth-value",
        }
        script = self.data_dir / "controlled_failure_updater.py"
        script.write_text(
            textwrap.dedent(
                f"""
                from pathlib import Path
                import sys

                workbook_path = Path({str(self.workbook_path)!r})
                workbook_path.write_bytes(b"this updater damaged the workbook")
                print("PORTAL_EMAIL={secrets['email']}", file=sys.stderr)
                print("PASSWORD={secrets['password']}", file=sys.stderr)
                print("access_token={secrets['token']}", file=sys.stderr)
                print("COOKIE: {secrets['cookie']}", file=sys.stderr)
                print("Authorization: Bearer {secrets['bearer']}", file=sys.stderr)
                print("Safe diagnostic: synthetic calendar failure", file=sys.stderr)
                raise SystemExit(7)
                """
            ).lstrip(),
            encoding="utf-8",
        )
        return script, secrets

    def write_zero_exit_corruption_updater(self) -> Path:
        """Create an updater that exits successfully but leaves invalid XLSX data."""

        script = self.data_dir / "controlled_zero_exit_corruption.py"
        script.write_text(
            textwrap.dedent(
                f"""
                from pathlib import Path

                workbook_path = Path({str(self.workbook_path)!r})
                workbook_path.write_bytes(b"not a valid Office Open XML archive")
                print("Updater claims success after writing spreadsheets", flush=True)
                """
            ).lstrip(),
            encoding="utf-8",
        )
        return script

    def write_slow_counting_updater(self, *, delay: float = 0.8) -> tuple[Path, Path]:
        """Create an updater whose append-only marker records each subprocess."""

        script = self.data_dir / "controlled_slow_counting_updater.py"
        marker = self.data_dir / "subprocess-starts.log"
        script.write_text(
            textwrap.dedent(
                f"""
                import os
                from pathlib import Path
                import time

                marker = Path({str(marker)!r})
                with marker.open("a", encoding="utf-8") as handle:
                    handle.write(f"{{os.getpid()}}\\n")
                    handle.flush()
                print("Reading calendar...", flush=True)
                time.sleep({delay!r})
                """
            ).lstrip(),
            encoding="utf-8",
        )
        return script, marker

    def assert_no_store(self, response) -> None:
        self.assertIn(
            "no-store",
            response.headers.get("Cache-Control", "").lower(),
            "refresh API responses must not be cached",
        )

    def wait_for_terminal_status(self, client, *, timeout: float = 8.0) -> dict:
        deadline = time.monotonic() + timeout
        latest: dict = {}
        while time.monotonic() < deadline:
            response = client.get("/api/refresh/status")
            self.assertEqual(response.status_code, 200)
            self.assert_no_store(response)
            latest = response.get_json()
            if latest.get("status") in {"success", "error"}:
                return latest
            time.sleep(0.025)
        self.fail(f"refresh did not finish within {timeout} seconds; latest={latest!r}")

    def test_start_is_nonblocking_double_start_is_rejected_and_reload_sees_running(self) -> None:
        script, marker = self.write_success_updater(delay=0.8)
        app = self.make_app(script)
        client = app.test_client()

        idle = client.get("/api/refresh/status")
        self.assertEqual(idle.status_code, 200)
        self.assert_no_store(idle)
        self.assertEqual(idle.get_json()["status"], "idle")

        before = time.monotonic()
        started = client.post("/api/refresh")
        elapsed = time.monotonic() - before

        self.assertEqual(started.status_code, 202)
        self.assert_no_store(started)
        self.assertLess(elapsed, 0.5, "POST must return before the updater finishes")
        started_payload = started.get_json()
        self.assertEqual(started_payload["status"], "running")
        self.assertTrue(started_payload["job_id"])
        self.assertTrue(started_payload["started_at"])

        duplicate = client.post("/api/refresh")
        self.assertEqual(duplicate.status_code, 409)
        self.assert_no_store(duplicate)
        self.assertEqual(duplicate.get_json()["status"], "running")
        self.assertIn("already", duplicate.get_json()["message"].lower())

        # A page reload must not reset in-memory job state or permit another job.
        self.assertEqual(client.get("/").status_code, 200)
        during_reload = client.get("/api/refresh/status").get_json()
        self.assertEqual(during_reload["status"], "running")
        self.assertEqual(during_reload["job_id"], started_payload["job_id"])

        finished = self.wait_for_terminal_status(client)
        self.assertEqual(finished["status"], "success")
        self.assertTrue(marker.exists())

    def test_success_uses_controlled_script_rereads_xlsx_and_persists_last_success(self) -> None:
        script, marker = self.write_success_updater()
        app = self.make_app(script)
        client = app.test_client()

        before = client.get("/api/gigs").get_json()
        self.assertEqual(len(before["rows"]), 1)
        self.assertEqual(before["summary"]["new"], 0)

        response = client.post("/api/refresh")
        self.assertEqual(response.status_code, 202)
        self.assert_no_store(response)
        finished = self.wait_for_terminal_status(client)

        self.assertEqual(finished["status"], "success")
        self.assertEqual(finished["new_count"], 1)
        self.assertTrue(finished["finished_at"])
        self.assertTrue(finished["last_success_at"])
        self.assertTrue(finished["last_success_display"])
        self.assertIn("success", finished["message"].lower())

        invocation = json.loads(marker.read_text(encoding="utf-8"))
        self.assertEqual(
            Path(invocation["executable"]).resolve(), Path(sys.executable).resolve()
        )
        self.assertEqual(
            [Path(value).resolve() for value in invocation["argv"]],
            [script.resolve()],
        )
        self.assertEqual(Path(invocation["cwd"]).resolve(), self.data_dir)

        after = client.get("/api/gigs").get_json()
        self.assertEqual(len(after["rows"]), 2)
        self.assertEqual(after["summary"]["new"], 1)
        new_rows = [row for row in after["rows"] if row["status"] == "NEW"]
        self.assertEqual([row["event_id"] for row in new_rows], ["new-2"])

        # The timestamp is stored on disk, not merely held by this Flask app.
        restarted_app = self.make_app(script)
        restarted_status = (
            restarted_app.test_client().get("/api/refresh/status").get_json()
        )
        self.assertEqual(restarted_status["last_success_at"], finished["last_success_at"])
        self.assertEqual(
            restarted_status["last_success_display"],
            finished["last_success_display"],
        )
        self.assertTrue((self.data_dir / "dashboard_state.json").is_file())

        restarted_client = restarted_app.test_client()
        restarted_snapshot = restarted_client.get("/api/gigs").get_json()
        self.assertEqual(
            restarted_snapshot["last_refreshed_iso"], finished["last_success_at"]
        )
        self.assertEqual(
            restarted_snapshot["last_refreshed"], finished["last_success_display"]
        )
        server_rendered_page = restarted_client.get("/")
        self.assertEqual(server_rendered_page.status_code, 200)
        self.assertIn(
            finished["last_success_display"], server_rendered_page.get_data(as_text=True)
        )

    def test_failure_restores_workbook_and_returns_only_sanitized_details(self) -> None:
        script, secrets = self.write_destructive_failure_updater()
        original_digest = digest(self.workbook_path)
        app = self.make_app(script)
        client = app.test_client()

        self.assertEqual(client.post("/api/refresh").status_code, 202)
        finished = self.wait_for_terminal_status(client)

        self.assertEqual(finished["status"], "error")
        self.assertIn("failed", finished["message"].lower())
        self.assertTrue(finished["details"])
        self.assertEqual(digest(self.workbook_path), original_digest)

        browser_visible_payload = json.dumps(finished)
        for label, secret in secrets.items():
            self.assertNotIn(
                secret,
                browser_visible_payload,
                f"{label} leaked through /api/refresh/status",
            )
        self.assertIn("synthetic calendar failure", finished["details"])

        # Failed output must not displace the last known-good dashboard data.
        snapshot = client.get("/api/gigs").get_json()
        self.assertEqual(len(snapshot["rows"]), 1)
        self.assertEqual(snapshot["rows"][0]["event_id"], "existing-1")

    def test_zero_exit_with_corrupt_xlsx_fails_strict_validation_and_rolls_back(self) -> None:
        script = self.write_zero_exit_corruption_updater()
        original_bytes = self.workbook_path.read_bytes()
        app = self.make_app(script)
        client = app.test_client()

        response = client.post("/api/refresh")
        self.assertEqual(response.status_code, 202)
        self.assert_no_store(response)
        finished = self.wait_for_terminal_status(client)

        self.assertEqual(finished["status"], "error")
        self.assertEqual(self.workbook_path.read_bytes(), original_bytes)
        self.assertIn("BadZipFile", finished["details"])
        self.assertNotIn("exited with code", finished["details"])
        snapshot = client.get("/api/gigs").get_json()
        self.assertEqual([row["event_id"] for row in snapshot["rows"]], ["existing-1"])

    def test_zero_exit_with_corrupt_csv_fails_validation_and_rolls_back_both_formats(self) -> None:
        existing_csv = self.data_dir / "US_Event_August_2026_Gigs.csv"
        existing_csv.write_text(
            "Date,Raw Calendar Listing,Status,Event ID\n"
            '2026-08-14,"4p WA|| Safeway 3265 (Chelan)",EXISTING,existing-1\n',
            encoding="utf-8",
        )
        original_xlsx = self.workbook_path.read_bytes()
        original_csv = existing_csv.read_bytes()
        script = self.data_dir / "controlled_zero_exit_csv_corruption.py"
        malformed_csv = "Date,Raw Calendar Listing\n2026-08-14\n"
        script.write_text(
            textwrap.dedent(
                f"""
                from pathlib import Path

                csv_path = Path({str(existing_csv)!r})
                csv_path.write_text({malformed_csv!r}, encoding="utf-8")
                print("Updater claims success after writing CSV export", flush=True)
                """
            ).lstrip(),
            encoding="utf-8",
        )
        app = self.make_app(script)
        client = app.test_client()

        response = client.post("/api/refresh")
        self.assertEqual(response.status_code, 202)
        self.assert_no_store(response)
        finished = self.wait_for_terminal_status(client)

        self.assertEqual(finished["status"], "error")
        self.assertIn("CSV validation failed", finished["details"])
        self.assertIn("malformed row", finished["details"])
        self.assertNotIn("exited with code", finished["details"])
        self.assertEqual(self.workbook_path.read_bytes(), original_xlsx)
        self.assertEqual(existing_csv.read_bytes(), original_csv)
        snapshot = client.get("/api/gigs").get_json()
        self.assertEqual([row["event_id"] for row in snapshot["rows"]], ["existing-1"])

    def test_failure_restores_existing_xlsx_and_csv_and_deletes_new_outputs(self) -> None:
        existing_csv = self.data_dir / "US_Event_August_2026_Gigs.csv"
        existing_csv.write_bytes(b"Date,Status,Event ID\n2026-08-14,EXISTING,existing-1\n")
        original_xlsx = self.workbook_path.read_bytes()
        original_csv = existing_csv.read_bytes()
        new_xlsx = self.data_dir / "US_Event_September_2026_Gigs.xlsx"
        new_csv = self.data_dir / "US_Event_September_2026_Gigs.csv"
        script = self.data_dir / "controlled_multi_output_failure.py"
        script.write_text(
            textwrap.dedent(
                f"""
                from pathlib import Path

                Path({str(self.workbook_path)!r}).write_bytes(b"changed existing xlsx")
                Path({str(existing_csv)!r}).write_bytes(b"changed existing csv")
                Path({str(new_xlsx)!r}).write_bytes(b"brand new xlsx")
                Path({str(new_csv)!r}).write_bytes(b"brand new csv")
                print("Synthetic failure after updating every export", flush=True)
                raise SystemExit(23)
                """
            ).lstrip(),
            encoding="utf-8",
        )
        app = self.make_app(script)
        client = app.test_client()

        self.assertEqual(client.post("/api/refresh").status_code, 202)
        finished = self.wait_for_terminal_status(client)

        self.assertEqual(finished["status"], "error")
        self.assertEqual(self.workbook_path.read_bytes(), original_xlsx)
        self.assertEqual(existing_csv.read_bytes(), original_csv)
        self.assertFalse(new_xlsx.exists())
        self.assertFalse(new_csv.exists())

    def test_independent_app_managers_share_one_process_lock(self) -> None:
        script, marker = self.write_slow_counting_updater()
        first_app = self.make_app(script)
        second_app = self.make_app(script)
        first_client = first_app.test_client()
        second_client = second_app.test_client()

        first = first_client.post("/api/refresh")
        self.assertEqual(first.status_code, 202)
        self.assert_no_store(first)
        second = second_client.post("/api/refresh")
        self.assertEqual(second.status_code, 409)
        self.assert_no_store(second)
        self.assertEqual(second.get_json()["status"], "running")
        self.assertIn("already", second.get_json()["message"].lower())

        self.assertEqual(self.wait_for_terminal_status(first_client)["status"], "success")
        starts = marker.read_text(encoding="utf-8").splitlines()
        self.assertEqual(len(starts), 1, "two updater subprocesses were started")

    def test_env_literal_is_redacted_even_without_a_secret_label(self) -> None:
        literal_secret = "literal-env-secret-7Qx9!"
        (self.data_dir / ".env").write_text(
            f'PORTAL_PASSWORD="{literal_secret}"\n', encoding="utf-8"
        )
        script = self.data_dir / "controlled_env_secret_failure.py"
        script.write_text(
            textwrap.dedent(
                f"""
                import sys

                print("Opaque upstream response contained {literal_secret}", file=sys.stderr)
                print("Safe diagnostic remains visible", file=sys.stderr)
                raise SystemExit(31)
                """
            ).lstrip(),
            encoding="utf-8",
        )
        app = self.make_app(script)
        client = app.test_client()

        self.assertEqual(client.post("/api/refresh").status_code, 202)
        finished = self.wait_for_terminal_status(client)

        self.assertEqual(finished["status"], "error")
        self.assertNotIn(literal_secret, json.dumps(finished))
        self.assertIn("[REDACTED]", finished["details"])
        self.assertIn("Safe diagnostic remains visible", finished["details"])

    def test_foreign_origin_is_rejected_without_starting_updater(self) -> None:
        script, marker = self.write_success_updater()
        app = self.make_app(script)
        client = app.test_client()

        response = client.post(
            "/api/refresh", headers={"Origin": "https://attacker.example"}
        )

        self.assertEqual(response.status_code, 403)
        self.assert_no_store(response)
        self.assertEqual(response.get_json()["status"], "error")
        self.assertIn("local dashboard", response.get_json()["message"].lower())
        self.assertFalse(marker.exists())
        status = client.get("/api/refresh/status").get_json()
        self.assertEqual(status["status"], "idle")

    def test_refresh_api_cannot_select_a_command_script_or_executable(self) -> None:
        script, marker = self.write_success_updater()
        rogue_script = self.data_dir / "rogue.py"
        rogue_marker = self.data_dir / "rogue-was-executed"
        rogue_script.write_text(
            f"from pathlib import Path\nPath({str(rogue_marker)!r}).touch()\n",
            encoding="utf-8",
        )
        app = self.make_app(script)
        client = app.test_client()

        response = client.post(
            "/api/refresh?script=/tmp/not-allowed.py&command=touch",
            json={
                "command": "touch rogue-was-executed",
                "executable": "/bin/sh",
                "filename": "main.py",
                "path": str(rogue_script),
                "script": str(rogue_script),
            },
        )
        self.assertEqual(response.status_code, 202)
        self.assertEqual(self.wait_for_terminal_status(client)["status"], "success")
        self.assertTrue(marker.exists(), "the server-configured updater was not run")
        self.assertFalse(rogue_marker.exists(), "request data selected an arbitrary command")

        # The two endpoints expose only their intended HTTP methods.
        self.assertEqual(client.get("/api/refresh").status_code, 405)
        self.assertEqual(client.post("/api/refresh/status").status_code, 405)


if __name__ == "__main__":
    unittest.main(verbosity=2)
