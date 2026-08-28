import csv
import calendar as calendar_module
import os
import re
import shutil
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from uuid import uuid4

from dotenv import load_dotenv
from playwright.sync_api import (
    Error as PlaywrightError,
    Playwright,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


ROOT = Path(__file__).resolve().parent
DEBUG_DIR = ROOT / "debug"
CSV_PATH = ROOT / "available_gigs.csv"
AUTH_STATE_PATH = ROOT / "auth_state.json"
BACKUP_DIR = ROOT / "backups"

LOGIN_URL = "https://portal.useventmanagement.com/login"
HOME_URL = "https://portal.useventmanagement.com/home"


@dataclass(frozen=True)
class MonthConfig:
    year: int
    month: int
    label: str
    xlsx_path: Path
    csv_path: Path
    sheet_name: str
    table_name: str

    @property
    def prefix(self) -> str:
        return f"{self.year:04d}-{self.month:02d}"

    @property
    def days_in_month(self) -> int:
        return calendar_module.monthrange(self.year, self.month)[1]


TARGET_MONTHS = (
    (2026, 8),
    (2026, 9),
    (2026, 10),
    (2026, 11),
)


def build_month_config(year: int, month: int) -> MonthConfig:
    month_name = calendar_module.month_name[month]
    if not month_name:
        raise ValueError(f"Invalid calendar month: {month}")
    label = f"{month_name} {year}"
    file_stem = f"US_Event_{month_name}_{year}_Gigs"
    return MonthConfig(
        year=year,
        month=month,
        label=label,
        xlsx_path=ROOT / f"{file_stem}.xlsx",
        csv_path=ROOT / f"{file_stem}.csv",
        sheet_name=f"{label} Gigs",
        table_name=f"{month_name}{year}Gigs",
    )


MONTH_CONFIGS = [build_month_config(year, month) for year, month in TARGET_MONTHS]

DAILY_GIG_HEADERS = [
    "Date",
    "Day",
    "Start Time",
    "Start Time Sort",
    "State",
    "Store / Account",
    "Store Number",
    "City",
    "Preferred Area",
    "Location Priority",
    "Raw Calendar Listing",
    "Calendar Color",
    "Event URL",
    "Status",
    "First Seen",
    "Last Seen",
    "Scraped At",
    "Event ID",
    "Role",
    "Event Time",
    "Pay",
    "Eligibility",
]

# Schema used before gig-detail (Role/Event Time/Pay/Eligibility) columns existed.
DAILY_GIG_HEADERS_V1 = DAILY_GIG_HEADERS[:18]

# Row fields populated by opening each gig's detail view.
DETAIL_FIELDS = ("role", "event_time", "pay", "eligibility")

LEGACY_GIG_HEADERS = [
    "Date",
    "Day",
    "Start Time",
    "Start Time Sort",
    "State",
    "Store / Account",
    "Store Number",
    "City",
    "Preferred Area",
    "Location Priority",
    "Raw Calendar Listing",
    "Calendar Color",
    "Event URL",
    "Scraped At",
]


def ensure_debug_dir() -> None:
    DEBUG_DIR.mkdir(exist_ok=True)


def load_credentials() -> tuple[str, str]:
    env_path = ROOT / ".env"
    load_dotenv(dotenv_path=env_path)

    email = os.getenv("US_EVENT_EMAIL", "").strip()
    password = os.getenv("US_EVENT_PASSWORD", "").strip()

    if not email or not password:
        raise RuntimeError(
            "Missing credentials. Add US_EVENT_EMAIL and US_EVENT_PASSWORD to the .env file in this project."
        )

    if "PUT_EMAIL_HERE" in email or "PUT_PASSWORD_HERE" in password:
        raise RuntimeError(
            "Credentials still contain placeholder values. Update the .env file with your real account details."
        )

    return email, password


def save_debug_screenshot(page, label: str) -> None:
    ensure_debug_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = DEBUG_DIR / f"{label}_{timestamp}.png"
    page.screenshot(path=str(file_path), full_page=True)
    print(f"Screenshot saved to: {file_path}")


def print_error_and_exit(message: str, page=None) -> None:
    print(f"ERROR: {message}")
    if page is not None:
        try:
            print(f"Current URL: {page.url}")
        except Exception:
            pass
    save_debug_screenshot(page, "error") if page is not None else None
    raise SystemExit(1)


def wait_for_selectable(locator, timeout: int = 20000):
    locator.wait_for(state="visible", timeout=timeout)
    return locator


def find_by_text(page, pattern: str, timeout: int = 15000):
    try:
        return page.get_by_text(re.compile(pattern, re.IGNORECASE), exact=False).first
    except Exception:
        return None


def find_first(page, selectors: list[str]):
    for selector in selectors:
        locator = page.locator(selector)
        if locator.count() > 0:
            return locator.first
    return None


def is_human_verification(page) -> bool:
    if page is None:
        return False

    verification_patterns = [
        "captcha",
        "verify you are human",
        "mfa",
        "two-factor",
        "two factor",
        "one-time code",
        "enter code",
        "security check",
        "verify identity",
    ]
    page_text = page.content().lower()
    return any(pattern in page_text for pattern in verification_patterns)


def detect_login_result(page) -> tuple[bool, str]:
    page_text = page.content().lower()
    login_failed_terms = [
        "incorrect email",
        "incorrect password",
        "invalid email",
        "invalid password",
        "wrong email",
        "wrong password",
        "email or password",
        "sign in failed",
        "authentication failed",
    ]

    if any(term in page_text for term in login_failed_terms):
        return False, "Invalid login credentials or authentication failure."

    if "forgot password" in page_text and page.url.lower().endswith("/login"):
        return False, "Login form is still visible."

    if is_human_verification(page):
        return False, "Human verification required (CAPTCHA/MFA/2FA)."

    if page.url.lower().startswith("https://portal.useventmanagement.com") and "/login" not in page.url.lower():
        return True, "Logged in successfully."

    return False, "Login status unknown."


def login(page) -> None:
    page.goto(LOGIN_URL, wait_until="domcontentloaded")

    # Some builds perform the saved-session redirect client-side just after the
    # initial document load. Give that redirect a short opportunity to finish.
    if "/login" in page.url.lower():
        try:
            page.wait_for_url(re.compile(r"^https://portal\.useventmanagement\.com/(?!login(?:[/?#]|$)).+"), timeout=5000)
        except Exception:
            pass

    # A valid saved session redirects /login to an authenticated route (normally
    # /home).  Do not inspect or wait for login controls after that redirect.
    if "/login" not in page.url.lower():
        # A stale saved session can pass the first redirect and then get bounced
        # back to /login moments later. Confirm we truly land on /home before
        # trusting the session; otherwise fall through to a fresh login.
        try:
            page.wait_for_url(HOME_URL, timeout=15000)
            print(f"Already authenticated; skipping login fields. Current URL: {page.url}")
            return
        except Exception:
            print(
                "Saved session appears stale (did not settle on /home); "
                f"performing a fresh login. Current URL: {page.url}"
            )
            page.goto(LOGIN_URL, wait_until="domcontentloaded")
            page.wait_for_timeout(1000)
            if "/login" not in page.url.lower():
                page.wait_for_url(HOME_URL, timeout=15000)
                print(f"Session recovered after reload. Current URL: {page.url}")
                return

    email_field = page.get_by_role("textbox", name=re.compile(r"email", re.I))
    password_field = page.get_by_role("textbox", name=re.compile(r"password", re.I))
    login_button = page.get_by_role("button", name=re.compile(r"log in|login", re.I))

    if email_field.count() == 0:
        email_field = page.locator('input[type="email"], input[name*="email" i], input[autocomplete*="email" i]').first
    if password_field.count() == 0:
        password_field = page.locator('input[type="password"], input[name*="password" i], input[autocomplete*="current-password" i]').first
    if login_button.count() == 0:
        login_button = page.locator('button:has-text("Log In"), button:has-text("LOGIN"), input[type="submit"]').first

    # Load credentials only when the browser is genuinely showing the login
    # form.  Existing authenticated sessions never enter the field logic.
    email, password = load_credentials()

    wait_for_selectable(email_field)
    wait_for_selectable(password_field)
    wait_for_selectable(login_button)

    email_field.fill(email)
    password_field.fill(password)
    login_button.click()

    # The SPA validates credentials asynchronously (button shows "Checking..").
    # Wait for the redirect away from /login rather than a network lull.
    try:
        page.wait_for_url(
            re.compile(r"^https://portal\.useventmanagement\.com/(?!login(?:[/?#]|$)).+"),
            timeout=45000,
        )
    except Exception:
        pass
    page.wait_for_load_state("networkidle", timeout=20000)

    success, message = detect_login_result(page)
    if success:
        print(message)
        return

    if is_human_verification(page):
        print("A CAPTCHA, MFA, or another human verification challenge appeared.")
        print("Please complete the required verification in the browser, then press Enter here to continue.")
        input("Press Enter after verification is complete: ")
        page.wait_for_load_state("networkidle", timeout=20000)
        success, message = detect_login_result(page)
        if success:
            print(message)
            return

    print_error_and_exit(f"Login failed. {message}", page)


def wait_for_dashboard(page) -> None:
    dashboard_patterns = [
        "dashboard",
        "available",
        "gigs",
        "assignments",
        "shifts",
        "jobs",
        "home",
    ]

    for text in dashboard_patterns:
        locator = find_by_text(page, text)
        if locator is not None:
            try:
                locator.wait_for(state="visible", timeout=20000)
                print(f"Dashboard-like element detected: {text}")
                return
            except Exception:
                pass

    page.wait_for_load_state("networkidle", timeout=20000)

    nav = page.locator('nav, [role="navigation"], header').first
    if nav.count() > 0:
        try:
            nav.wait_for(state="visible", timeout=20000)
            print("Navigation bar detected.")
            return
        except Exception:
            pass

    raise RuntimeError("Dashboard did not appear after login.")


def clean_text(value: str) -> str:
    if value is None:
        return ""
    value = value.replace("\xa0", " ").replace("\n", " ").strip()
    value = re.sub(r"\s+", " ", value)
    return value


def parse_date_and_day(value: str) -> tuple[str, str]:
    value = clean_text(value)
    if not value:
        return "", ""

    candidates = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%b %d, %Y",
        "%B %d, %Y",
        "%a %b %d, %Y",
        "%a, %b %d",
        "%A, %b %d",
        "%b %d",
        "%m-%d-%Y",
        "%m-%d-%y",
        "%d %b %Y",
    ]

    for fmt in candidates:
        try:
            dt = datetime.strptime(value, fmt)
            return dt.strftime("%Y-%m-%d"), dt.strftime("%A")
        except ValueError:
            pass

    # Fallback: extract any date-like fragment from strings such as 'Mon, Aug 12'
    match = re.search(r"(Mon|Tue|Wed|Thu|Fri|Sat|Sun)[a-zA-Z]*[, ]+[A-Z][a-z]+\s+\d{1,2}(?:,\s*\d{4})?", value)
    if match:
        parsed = match.group(0)
        try:
            dt = datetime.strptime(parsed, "%a, %b %d")
            return dt.strftime("%Y-%m-%d"), dt.strftime("%A")
        except ValueError:
            pass

    return value, ""


def parse_time_range(text: str) -> tuple[str, str]:
    clean = clean_text(text)
    if not clean:
        return "", ""

    times = re.findall(r"\d{1,2}:\d{2}\s?(?:AM|PM|am|pm)?", clean)
    if len(times) >= 2:
        return times[0], times[1]
    if len(times) == 1:
        return times[0], ""
    return "", ""


def parse_store_and_location(text: str) -> tuple[str, str, str, str, str, str]:
    cleaned = clean_text(text)
    if not cleaned:
        return "", "", "", "", "", ""

    store_name = ""
    store_number = ""
    city = ""
    state = ""
    address = ""

    # Try to capture a store number near a # or number label.
    store_number_match = re.search(r"(?:Store|Location|#)\s*[:#-]?\s*(\d{2,8})", cleaned, re.I)
    if store_number_match:
        store_number = store_number_match.group(1)

    # Look for a U.S. city/state pattern like 'Dallas, TX'
    city_state_match = re.search(r"([A-Za-z. ]+),\s*([A-Z]{2})\b", cleaned)
    if city_state_match:
        city = city_state_match.group(1).strip()
        state = city_state_match.group(2).strip()

    # Find persisted street address if present.
    address_match = re.search(r"\d+\s+[A-Za-z0-9. ]+?(?:Street|St|Road|Rd|Ave|Avenue|Boulevard|Blvd|Drive|Dr|Lane|Ln|Way|Parkway|Pkwy|Court|Ct|Place|Pl)\b.*?", cleaned, re.I)
    if address_match:
        address = address_match.group(0).strip()

    # Store name heuristics: often appears before city or at start of row.
    if re.search(r"\b(Store|Location|Account|Branch)\b", cleaned, re.I):
        store_name = re.split(r"\b(?:Store|Location|Account|Branch)\b", cleaned, flags=re.I, maxsplit=1)[0].strip(" -:")
    elif " at " in cleaned.lower():
        store_name = cleaned.lower().split(" at ", 1)[0].strip(" -:")

    if not store_name:
        # Common fallback: take the leading chunk before a city/state or date-like segment.
        slug = re.split(r"\s+-\s+|\s+(?:\d{1,2}/\d{1,2}/\d{2,4}|[A-Z][a-z]+,\s*[A-Z]{2})", cleaned, maxsplit=1)[0]
        if len(slug) < 80:
            store_name = slug.strip(" -:")

    return store_name, store_number, city, state, address, cleaned


def normalize_row_values(row: dict) -> dict:
    normalized = {
        "date": "",
        "day": "",
        "start_time": "",
        "end_time": "",
        "store": "",
        "store_number": "",
        "city": "",
        "state": "",
        "address": "",
        "pay": "",
        "brand": "",
        "status": "",
    }

    for key, value in row.items():
        raw = clean_text(str(value))
        lowered = raw.lower()

        if key in {"date", "day"}:
            continue

        if "date" in key.lower() and raw:
            normalized["date"], normalized["day"] = parse_date_and_day(raw)
        elif "day" in key.lower() and raw:
            normalized["day"] = raw

        if "time" in key.lower() or "start" in key.lower() or "end" in key.lower():
            if "start" in key.lower() and raw:
                normalized["start_time"] = raw
            if "end" in key.lower() and raw:
                normalized["end_time"] = raw
            if "time" in key.lower() and raw:
                s_time, e_time = parse_time_range(raw)
                if s_time:
                    normalized["start_time"] = s_time
                if e_time:
                    normalized["end_time"] = e_time

        if any(term in key.lower() for term in ["store", "location", "account", "branch"]):
            if raw:
                store_name, store_number, city, state, address, _ = parse_store_and_location(raw)
                normalized["store"] = store_name or raw
                normalized["store_number"] = store_number
                normalized["city"] = city
                normalized["state"] = state
                normalized["address"] = address

        if "pay" in key.lower() or "rate" in key.lower() or "amount" in key.lower():
            normalized["pay"] = raw
        if "brand" in key.lower() or "product" in key.lower() or "event" in key.lower():
            normalized["brand"] = raw
        if "status" in key.lower():
            normalized["status"] = raw

    if not normalized["date"]:
        for candidate in [row.get("date"), row.get("Date"), row.get("event_date"), row.get("shift_date")]:
            if candidate:
                normalized["date"], normalized["day"] = parse_date_and_day(str(candidate))
                break

    if not normalized["day"] and normalized["date"]:
        try:
            normalized["day"] = datetime.strptime(normalized["date"], "%Y-%m-%d").strftime("%A")
        except ValueError:
            pass

    if not normalized["start_time"] and not normalized["end_time"]:
        combined_text = " ".join(str(v) for v in row.values() if v)
        s_time, e_time = parse_time_range(combined_text)
        normalized["start_time"] = s_time
        normalized["end_time"] = e_time

    return normalized


def collect_candidate_rows(page) -> list[str]:
    candidates = []
    selectors = [
        "table tr",
        "[role='row']",
        "li",
        "article",
        "[data-testid*='gig']",
        "[data-testid*='shift']",
        "[data-testid*='assignment']",
        "[data-testid*='job']",
        ".MuiCard-root",
        "[class*='card']",
    ]

    for selector in selectors:
        loc = page.locator(selector)
        for i in range(min(loc.count(), 200)):
            text = clean_text(loc.nth(i).inner_text())
            if not text:
                continue
            if len(text) < 10:
                continue
            if re.search(r"(Mon|Tue|Wed|Thu|Fri|Sat|Sun|\d{1,2}/\d{1,2}/\d{2,4}|\d{1,2}:\d{2})", text, re.I):
                candidates.append(text)
    return candidates


def row_text_candidates_to_rows(raw_rows: list[str]) -> list[dict]:
    rows = []
    for raw_text in raw_rows:
        values = [part.strip() for part in raw_text.split("\n") if part.strip()]
        if len(values) <= 1:
            continue
        row = {f"col_{idx}": value for idx, value in enumerate(values[:12])}
        rows.append(normalize_row_values(row))
    return rows


def export_csv(rows: list[dict], output_path: Path) -> None:
    fieldnames = [
        "date",
        "day",
        "start_time",
        "end_time",
        "store",
        "store_number",
        "city",
        "state",
        "address",
        "pay",
        "brand",
        "status",
    ]

    with output_path.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def click_top_header_calendar(page) -> None:
    discovery = page.evaluate(
        """
        () => {
          const HEADER_BOTTOM = 160;
          const targetAttribute = 'data-us-event-top-calendar-target';
          document.querySelectorAll(`[${targetAttribute}]`).forEach(el => el.removeAttribute(targetAttribute));

          const value = (el, name) => el.getAttribute(name) || '';
          const classValue = el => value(el, 'class');
          const rectObject = rect => ({
            x: Math.round(rect.x),
            y: Math.round(rect.y),
            top: Math.round(rect.top),
            right: Math.round(rect.right),
            bottom: Math.round(rect.bottom),
            left: Math.round(rect.left),
            width: Math.round(rect.width),
            height: Math.round(rect.height),
          });
          const isVisibleInTopHeader = el => {
            const rect = el.getBoundingClientRect();
            const style = getComputedStyle(el);
            return style.display !== 'none' &&
              style.visibility !== 'hidden' &&
              Number(style.opacity || '1') > 0 &&
              rect.width > 0 && rect.height > 0 &&
              rect.right > 0 && rect.left < innerWidth &&
              rect.bottom > 0 && rect.top < HEADER_BOTTOM &&
              rect.height <= HEADER_BOTTOM;
          };
          const fields = el => {
            const descendantClasses = [...el.querySelectorAll('[class]')]
              .map(node => classValue(node)).join(' ');
            const descendantText = (el.innerText || el.textContent || '')
              .replace(/\\s+/g, ' ').trim();
            return {
              href: value(el, 'href') || el.href || '',
              className: classValue(el),
              id: el.id || '',
              title: value(el, 'title'),
              ariaLabel: value(el, 'aria-label'),
              descendantClasses,
              descendantText,
              outerHTML: el.outerHTML || '',
            };
          };
          const containsCalendar = data => [
            data.href,
            data.className,
            data.id,
            data.title,
            data.ariaLabel,
            data.descendantClasses,
            data.descendantText,
            data.outerHTML,
          ].some(item => String(item).toLowerCase().includes('calendar'));
          const isSidebarElement = el => Boolean(el.closest(
            'aside, [class*="sidebar" i], [id*="sidebar" i], [class*="sidenav" i], [id*="sidenav" i], [class*="drawer" i], [id*="drawer" i]'
          ));

          const matches = [...document.querySelectorAll('*')]
            .filter(isVisibleInTopHeader)
            .map(el => ({ el, data: fields(el) }))
            .filter(item => containsCalendar(item.data));

          const candidates = matches.map(({ el, data }) => ({
            tag: el.tagName.toLowerCase(),
            href: data.href,
            id: data.id,
            className: data.className,
            title: data.title,
            ariaLabel: data.ariaLabel,
            boundingBox: rectObject(el.getBoundingClientRect()),
            outerHTML: data.outerHTML,
          }));

          // Resolve semantic calendar graphics to their actual clickable parent.
          // No hamburger relationship, sibling order, or click coordinates are used.
          const clickableChoices = [];
          const seen = new Set();
          for (const { el } of matches) {
            const clickable = el.matches('a, button') ? el : el.closest('a, button');
            if (!clickable || seen.has(clickable) || !isVisibleInTopHeader(clickable) || isSidebarElement(clickable)) {
              continue;
            }
            const clickableData = fields(clickable);
            if (!containsCalendar(clickableData)) {
              continue;
            }
            const semanticDescendants = [...clickable.querySelectorAll('*')]
              .filter(node => containsCalendar(fields(node)));
            const directMetadata = [
              clickableData.href,
              clickableData.className,
              clickableData.id,
              clickableData.title,
              clickableData.ariaLabel,
            ].some(item => String(item).toLowerCase().includes('calendar'));
            const descendantMetadata = semanticDescendants.some(node => {
              const data = fields(node);
              return [data.className, data.id, data.title, data.ariaLabel]
                .some(item => String(item).toLowerCase().includes('calendar'));
            });
            const descendantIconMetadata = semanticDescendants.some(node => {
              if (!node.matches('i, svg, use, path, img, span, [class*="icon" i]')) {
                return false;
              }
              const data = fields(node);
              return [data.className, data.id, data.title, data.ariaLabel, data.outerHTML]
                .some(item => String(item).toLowerCase().includes('calendar'));
            });
            const hrefIsCalendar = clickableData.href.toLowerCase().includes('calendar');
            const signature = `${clickableData.id} ${clickableData.className} ${clickableData.title} ${clickableData.ariaLabel}`.toLowerCase();
            if (/hamburger|menu-toggle|menu_button|menubutton/.test(signature)) {
              continue;
            }
            seen.add(clickable);
            clickableChoices.push({
              el: clickable,
              data: clickableData,
              score: (descendantIconMetadata ? 200 : 0) + (hrefIsCalendar ? 100 : 0) +
                (directMetadata ? 50 : 0) + (descendantMetadata ? 25 : 0),
            });
          }

          clickableChoices.sort((a, b) => b.score - a.score);
          const chosen = clickableChoices[0];
          if (!chosen) {
            return { candidates, chosen: null };
          }

          const chosenHTML = chosen.el.outerHTML;
          chosen.el.setAttribute(targetAttribute, 'true');
          return {
            candidates,
            chosen: {
              href: chosen.data.href,
              outerHTML: chosenHTML,
              boundingBox: rectObject(chosen.el.getBoundingClientRect()),
            },
          };
        }
        """
    )

    print("Visible DOM candidates containing 'calendar' in the top header:")
    if not discovery["candidates"]:
        print("  (none)")
    for index, candidate in enumerate(discovery["candidates"], start=1):
        print(f"Candidate {index}:")
        print(f"  tag: {candidate['tag']}")
        print(f"  href: {candidate['href']}")
        print(f"  id: {candidate['id']}")
        print(f"  class: {candidate['className']}")
        print(f"  title: {candidate['title']}")
        print(f"  aria-label: {candidate['ariaLabel']}")
        print(f"  bounding box: {candidate['boundingBox']}")
        print(f"  outerHTML: {candidate['outerHTML']}")

    chosen = discovery["chosen"]
    if chosen is None:
        raise RuntimeError(
            "No visible clickable <a> or <button> with calendar DOM semantics was found in the top header."
        )

    print("FOUND TOP CALENDAR:")
    print(chosen["outerHTML"])
    print("Calendar href:", chosen["href"])

    old_url = page.url
    print("URL before click:", old_url)
    print("About to click TOP CALENDAR ICON")

    calendar_locator = page.locator('[data-us-event-top-calendar-target="true"]')
    if calendar_locator.count() != 1:
        raise RuntimeError(
            f"Expected exactly one verified top-calendar locator, found {calendar_locator.count()}."
        )

    try:
        calendar_locator.click(timeout=10000)
    except Exception as normal_click_error:
        print(f"Normal Playwright click failed: {normal_click_error}")
        try:
            calendar_locator.click(force=True, timeout=10000)
        except Exception as force_click_error:
            print(f"Forced Playwright click failed: {force_click_error}")
            print("Trying element.click() on the same verified calendar element.")
            calendar_locator.evaluate("element => element.click()")

    if page.url == old_url:
        try:
            page.wait_for_url(lambda url: str(url) != old_url, timeout=5000)
        except Exception:
            pass

    print("URL after click:", page.url)
    page.wait_for_timeout(15000)
    print("Calendar view verification wait complete; starting read-only DOM inspection.")


MONTH_TITLE_PATTERN = re.compile(r"\b([A-Za-z]+)\s+(20\d{2})\b")


def month_key_from_title(value: str) -> str:
    match = MONTH_TITLE_PATTERN.search(clean_text(value))
    if not match:
        return ""
    for format_string in ("%B %Y", "%b %Y"):
        try:
            return datetime.strptime(match.group(0), format_string).strftime("%Y-%m")
        except ValueError:
            pass
    return ""


def month_counts_from_dates(values: list[str], distinct_dates: bool) -> dict[str, int]:
    normalized_dates = []
    for value in values:
        match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", clean_text(str(value or "")))
        if not match:
            continue
        try:
            normalized_dates.append(date.fromisoformat(match.group(1)).isoformat())
        except ValueError:
            continue
    if distinct_dates:
        normalized_dates = sorted(set(normalized_dates))
    counts = Counter(item[:7] for item in normalized_dates)
    return dict(sorted(counts.items()))


def unique_dominant_month(month_counts: dict[str, int]) -> tuple[str, int]:
    if not month_counts:
        return "", 0
    largest_count = max(month_counts.values())
    winners = [month for month, count in month_counts.items() if count == largest_count]
    return (winners[0], largest_count) if len(winners) == 1 else ("", largest_count)


def month_label_from_key(month_key: str) -> str:
    try:
        return datetime.strptime(month_key, "%Y-%m").strftime("%B %Y")
    except ValueError:
        return month_key or "<unknown>"


def expected_month_dates(config: MonthConfig) -> set[str]:
    return {
        f"{config.prefix}-{day_number:02d}"
        for day_number in range(1, config.days_in_month + 1)
    }


def sparse_event_fallback_support(
    state: dict, config: MonthConfig
) -> tuple[bool, str]:
    """Return whether event metadata safely corroborates a sparse day grid.

    Event dates never stand alone as month evidence. This fallback is limited
    to a genuinely sparse/nonstandard grid with some target-month day metadata;
    a mostly rendered standard grid must finish rendering instead.
    """
    day_cell_dates = sorted(
        {
            match.group(1)
            for value in state.get("dayCellDates", [])
            if (match := re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", str(value or "")))
        }
    )
    if not day_cell_dates:
        return False, "no day-cell metadata was available"
    if state.get("isDayGrid"):
        return False, "the FullCalendar day-grid was only partially rendered"
    if not state.get("isSparseView"):
        return False, "the DOM did not identify a recognized sparse/list calendar view"
    if len(day_cell_dates) >= 14:
        return False, "the standard month grid was only partially rendered"

    title_month = month_key_from_title(state.get("title", ""))
    if title_month and title_month != config.prefix:
        return False, "the visible toolbar identifies a different month"

    day_counts = month_counts_from_dates(day_cell_dates, distinct_dates=True)
    target_day_count = day_counts.get(config.prefix, 0)
    if not target_day_count or target_day_count < max(day_counts.values(), default=0):
        return False, "the sparse day grid does not support the target month"

    event_counts = month_counts_from_dates(
        state.get("eventDates", []), distinct_dates=False
    )
    dominant_event_month, _ = unique_dominant_month(event_counts)
    if dominant_event_month != config.prefix:
        return False, "event metadata does not uniquely support the target month"
    if not state.get("root"):
        return False, "no verified calendar root was found"
    if state.get("undatedEventCount", 0):
        return False, "one or more FullCalendar events had no authoritative date"
    return True, "target day cells and unique event metadata agree"


def resolve_calendar_month_evidence(
    title: str, day_cell_dates: list[str], event_dates: list[str]
) -> dict:
    title_month = month_key_from_title(title)
    day_month_counts = month_counts_from_dates(day_cell_dates, distinct_dates=True)
    event_month_counts = month_counts_from_dates(event_dates, distinct_dates=False)
    dominant_day_month, dominant_day_count = unique_dominant_month(day_month_counts)
    dominant_event_month, dominant_event_count = unique_dominant_month(event_month_counts)

    # Fourteen distinct dates prevents a partially rendered grid from being
    # mistaken for the active month. Standard FullCalendar month views provide
    # every day in the month plus adjacent-month spillover cells.
    day_evidence_is_strong = bool(dominant_day_month and dominant_day_count >= 14)
    if title_month:
        detected_month = title_month
        method = "visible toolbar title"
    elif dominant_day_month:
        detected_month = dominant_day_month
        method = "day-cell metadata"
    else:
        detected_month = ""
        method = "insufficient evidence"

    return {
        "titleMonth": title_month,
        "dayMonthCounts": day_month_counts,
        "dominantDayMonth": dominant_day_month,
        "dominantDayCount": dominant_day_count,
        "dayEvidenceIsStrong": day_evidence_is_strong,
        "eventMonthCounts": event_month_counts,
        "dominantEventMonth": dominant_event_month,
        "dominantEventCount": dominant_event_count,
        "detectedMonth": detected_month,
        "method": method,
    }


def read_calendar_month_state(page) -> dict:
    state = page.evaluate(
        """
        () => {
          const elements = [...document.querySelectorAll('*')];
          const clean = value => String(value || '').replace(/\\s+/g, ' ').trim();
          const datedElements = elements.filter(el => [...el.attributes]
            .some(attr => attr.name.toLowerCase() === 'data-date' && /^20\\d{2}-\\d{2}-\\d{2}$/.test(attr.value)));
          const visible = el => {
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
          };
          const titleSelectors = [
            '.fc-toolbar-title',
            '.fc-header-toolbar .fc-toolbar-title',
            '.fc-header-toolbar [role="heading"]',
            '.fc-header-toolbar h1, .fc-header-toolbar h2, .fc-header-toolbar h3',
            '[class*="calendar" i] [role="heading"]',
          ];
          const titleAttempts = titleSelectors.map(selector => {
            const nodes = [...document.querySelectorAll(selector)];
            const preferred = nodes.find(el => visible(el) && clean(el.textContent)) ||
              nodes.find(visible) || null;
            return {
              selector,
              foundCount: nodes.length,
              text: clean(preferred?.textContent),
            };
          });
          const titleElements = [...new Set(titleSelectors.flatMap(selector =>
            [...document.querySelectorAll(selector)]
          ))];
          const title = titleElements.find(el => visible(el) && clean(el.textContent)) ||
            titleElements.find(visible) || null;
          const titleAnchor = title;

          let calendarRoot = null;
          if (titleAnchor) {
            let node = titleAnchor.parentElement;
            while (node && node !== document.body) {
              const dates = new Set([...node.querySelectorAll('[data-date]')]
                .filter(visible)
                .map(el => el.getAttribute('data-date'))
                .filter(value => /^20\\d{2}-\\d{2}-\\d{2}$/.test(value || '')));
              if (dates.size >= 28) {
                calendarRoot = node;
                break;
              }
              node = node.parentElement;
            }
          }
          if (!calendarRoot && datedElements.length) {
            const semanticRoot = el => {
              const tag = el.tagName.toLowerCase();
              const classTokens = String(el.getAttribute('class') || '').split(/\\s+/);
              const signature = `${el.id || ''} ${el.getAttribute('class') || ''}`;
              return tag.includes('calendar') || classTokens.includes('fc') ||
                /full-calendar|calendar-container|calendar-wrapper/i.test(signature);
            };
            const rootCandidates = new Map();
            for (const datedElement of datedElements.filter(visible)) {
              let node = datedElement;
              while (node && node !== document.body) {
                if (semanticRoot(node) && visible(node)) {
                  const dates = new Set([...node.querySelectorAll('[data-date]')]
                    .filter(visible)
                    .map(el => el.getAttribute('data-date'))
                    .filter(value => /^20\\d{2}-\\d{2}-\\d{2}$/.test(value || '')));
                  rootCandidates.set(node, dates.size);
                }
                node = node.parentElement;
              }
            }
            calendarRoot = [...rootCandidates.entries()]
              .sort((a, b) => b[1] - a[1] ||
                a[0].querySelectorAll('*').length - b[0].querySelectorAll('*').length)[0]?.[0] || null;
          }

          document.querySelectorAll('[data-us-event-calendar-root]').forEach(el =>
            el.removeAttribute('data-us-event-calendar-root'));
          if (calendarRoot) calendarRoot.setAttribute('data-us-event-calendar-root', 'true');

          const scope = calendarRoot || document;
          const isDayGrid = Boolean(
            scope.matches?.('.fc-daygrid, .fc-dayGridMonth-view') ||
            scope.querySelector('.fc-daygrid, .fc-daygrid-day, .fc-dayGridMonth-view')
          );
          const isSparseView = Boolean(
            scope.matches?.('.fc-list, [class*="list-view" i], [class*="listMonth" i]') ||
            scope.querySelector('.fc-list, [class*="list-view" i], [class*="listMonth" i]')
          );
          const dayCellDates = [...new Set([...scope.querySelectorAll('[data-date]')]
            .filter(visible)
            .map(el => el.getAttribute('data-date') || '')
            .filter(value => /^20\\d{2}-\\d{2}-\\d{2}$/.test(value)))].sort();
          const monthCounts = {};
          for (const value of dayCellDates) {
            const match = value.match(/^(20\\d{2}-\\d{2})-\\d{2}$/);
            if (match) monthCounts[match[1]] = (monthCounts[match[1]] || 0) + 1;
          }
          const dateFromValue = value => {
            if (!value) return '';
            try {
              const serialized = typeof value.toISOString === 'function'
                ? value.toISOString()
                : clean(value);
              const match = serialized.match(/(20\\d{2}-\\d{2}-\\d{2})/);
              return match ? match[1] : '';
            } catch (_) {
              return '';
            }
          };
          const eventDates = [];
          let undatedEventCount = 0;
          const seenEventKeys = new Set();
          const eventElements = [...scope.querySelectorAll('*')].filter(el =>
            Boolean(el.fcSeg) && el.fcSeg?.isStart !== false &&
            !String(el.getAttribute('class') || '').toLowerCase().includes('event-mirror')
          );
          for (const el of eventElements) {
            const eventRange = el.fcSeg?.eventRange || {};
            const definition = eventRange.def || {};
            const instance = eventRange.instance || {};
            const eventDate = el.closest('[data-date]')?.getAttribute('data-date') ||
              dateFromValue(eventRange.range?.start) ||
              dateFromValue(instance.range?.start);
            const identity = clean(instance.instanceId || definition.publicId || definition.defId ||
              definition.title || el.textContent);
            const eventKey = `${identity}::${eventDate}`;
            if (seenEventKeys.has(eventKey)) continue;
            seenEventKeys.add(eventKey);
            if (eventDate) eventDates.push(eventDate);
            else undatedEventCount += 1;
          }
          return {
            title: clean(title?.textContent),
            titleAttempts,
            monthCounts,
            dayCellDates,
            eventDates,
            undatedEventCount,
            isDayGrid,
            isSparseView,
            root: calendarRoot ? {
              tag: calendarRoot.tagName.toLowerCase(),
              id: calendarRoot.id || '',
              className: calendarRoot.getAttribute('class') || '',
            } : null,
          };
        }
        """
    )

    evidence = resolve_calendar_month_evidence(
        state.get("title", ""),
        state.get("dayCellDates", []),
        state.get("eventDates", []),
    )
    state.update(evidence)
    state["monthCounts"] = evidence["dayMonthCounts"]
    state["monthKey"] = evidence["detectedMonth"]
    return state


def validate_calendar_month_state(state: dict, config: MonthConfig) -> dict:
    evidence = resolve_calendar_month_evidence(
        state.get("title", ""),
        state.get("dayCellDates", []),
        state.get("eventDates", []),
    )
    state.update(evidence)

    print(f"Expected month: {config.label}")
    diagnostic_attempts = []
    for index, attempt in enumerate(state.get("titleAttempts", [])):
        if index < 2 or attempt.get("text"):
            diagnostic_attempts.append(attempt)
        if attempt.get("text"):
            break
    for attempt in diagnostic_attempts:
        title_text = attempt.get("text", "") or "<empty>"
        print(f"Month title selector tried: {attempt.get('selector', '')}")
        print(f"Month title text: {title_text}")

    toolbar_title = clean_text(state.get("title", ""))
    print(f"Toolbar title: {toolbar_title or '<empty>'}")
    if not evidence["titleMonth"]:
        print(
            "Toolbar title unavailable or empty. "
            "Falling back to calendar metadata validation."
        )

    print("Calendar day-cell month counts:")
    if evidence["dayMonthCounts"]:
        for month_key, count in evidence["dayMonthCounts"].items():
            print(f"{month_key}: {count}")
    else:
        print("(none)")

    dominant_day_month = evidence["dominantDayMonth"]
    if dominant_day_month:
        print(f"Dominant calendar month: {month_label_from_key(dominant_day_month)}")
    elif evidence["dayMonthCounts"]:
        print("Dominant calendar month: <ambiguous>")
    else:
        print("Dominant calendar month: <unavailable>")

    title_month = evidence["titleMonth"]
    if title_month and title_month != config.prefix:
        raise RuntimeError(
            f"Expected {config.label}, but the visible calendar toolbar shows "
            f"{month_label_from_key(title_month)}."
        )

    if evidence["dayEvidenceIsStrong"] and dominant_day_month != config.prefix:
        if title_month:
            raise RuntimeError(
                f"Calendar evidence conflicts: the toolbar shows {config.label}, but "
                f"the dominant day-cell month is {month_label_from_key(dominant_day_month)}."
            )
        raise RuntimeError(
            f"Expected {config.label}, but the dominant day-cell month is "
            f"{month_label_from_key(dominant_day_month)}."
        )

    observed_dates = {
        match.group(1)
        for value in state.get("dayCellDates", [])
        if (match := re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", str(value or "")))
    }
    missing_target_dates = sorted(expected_month_dates(config) - observed_dates)
    if not missing_target_dates:
        if not title_month and dominant_day_month != config.prefix:
            raise RuntimeError(
                f"Could not validate {config.label}: the complete target dates were "
                "present, but the visible day-cell grid did not have a unique "
                "target-month majority."
            )
        validation_method = (
            "visible toolbar title" if title_month else "day-cell metadata"
        )
        state["targetGridComplete"] = True
    else:
        fallback_supported, fallback_reason = sparse_event_fallback_support(
            state, config
        )
        if not fallback_supported:
            missing_preview = ", ".join(missing_target_dates[:5])
            if len(missing_target_dates) > 5:
                missing_preview += ", ..."
            raise RuntimeError(
                f"Could not safely validate {config.label}: {fallback_reason}. "
                f"The calendar DOM is missing {len(missing_target_dates)} target "
                f"day cells ({missing_preview}); refusing to scrape a partial month."
            )
        print(
            "Standard day-cell grid unavailable. Using FullCalendar event metadata "
            "only as supporting evidence for the sparse target-month grid."
        )
        print("Calendar event-metadata month counts:")
        for month_key, count in evidence["eventMonthCounts"].items():
            print(f"{month_key}: {count}")
        validation_method = "sparse day cells supported by FullCalendar event metadata"
        state["targetGridComplete"] = False

    print(
        f"Validated displayed month from {validation_method}: {config.label}"
    )
    print("Month validation PASSED")
    state["validatedBy"] = validation_method
    state["monthKey"] = config.prefix
    return state


def mark_month_navigation_control(
    page, direction: str, current_key: str
) -> dict:
    return page.evaluate(
        """
        payload => {
          const direction = payload.direction;
          const currentKey = payload.currentKey;
          const clean = value => String(value || '').replace(/\\s+/g, ' ').trim();
          const explicitClickableSelector = [
            'button',
            'a',
            '[role="button"]',
            '[role="link"]',
            '[onclick]',
            '[ng-click]',
            '[data-ng-click]',
            '[tabindex]',
            'input[type="button"]',
            'input[type="submit"]',
          ].join(', ');
          const iconSelector = [
            'svg',
            'use',
            'i',
            'mat-icon',
            '[data-icon]',
            '[class*="icon" i]',
          ].join(', ');
          const visible = el => {
            if (!el || !(el instanceof Element)) return false;
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.display !== 'none' && style.visibility !== 'hidden' &&
              Number(style.opacity || '1') !== 0 && rect.width > 0 && rect.height > 0;
          };
          const enabled = el => !el.matches(':disabled') &&
            el.getAttribute('aria-disabled') !== 'true' &&
            !String(el.getAttribute('class') || '').toLowerCase().includes('disabled');
          const boundingBox = el => {
            const rect = el.getBoundingClientRect();
            return {
              x: Math.round(rect.x * 10) / 10,
              y: Math.round(rect.y * 10) / 10,
              width: Math.round(rect.width * 10) / 10,
              height: Math.round(rect.height * 10) / 10,
            };
          };
          const dataAttributes = el => Object.fromEntries([...el.attributes]
            .filter(attr => attr.name.toLowerCase().startsWith('data-'))
            .map(attr => [attr.name, attr.value]));
          const attributeText = el => [...el.attributes]
            .map(attr => `${attr.name}=${attr.value}`).join(' ');
          const visibleText = el => clean(el.innerText || el.textContent);
          const monthNames = [
            ['January', 'Jan'], ['February', 'Feb'], ['March', 'Mar'],
            ['April', 'Apr'], ['May', 'May'], ['June', 'Jun'],
            ['July', 'Jul'], ['August', 'Aug'], ['September', 'Sep'],
            ['October', 'Oct'], ['November', 'Nov'], ['December', 'Dec'],
          ];
          const keyMatch = /^(20\\d{2})-(0[1-9]|1[0-2])$/.exec(currentKey);
          if (!keyMatch) {
            return {
              count: 0,
              reason: `Invalid current month key: ${currentKey}`,
              monthLabel: null,
              toolbar: null,
              candidates: [],
              selected: null,
            };
          }
          const year = keyMatch[1];
          const [fullMonth, shortMonth] = monthNames[Number(keyMatch[2]) - 1];
          const monthLabelPattern = new RegExp(
            `^(?:${fullMonth}|${shortMonth})\\\\s*,?\\\\s*${year}$`, 'i'
          );
          const verifiedCalendarRoot = document.querySelector(
            '[data-us-event-calendar-root="true"]'
          );

          document.querySelectorAll('[data-us-event-month-nav-target]').forEach(el =>
            el.removeAttribute('data-us-event-month-nav-target'));

          const allElements = [...document.querySelectorAll('*')];
          const labelCandidates = allElements.filter(visible).filter(el => {
            const text = visibleText(el);
            const aria = clean(el.getAttribute('aria-label'));
            if (!monthLabelPattern.test(text) && !monthLabelPattern.test(aria)) return false;
            return ![...el.children].some(child => visible(child) &&
              monthLabelPattern.test(visibleText(child)));
          });

          const lowestCommonAncestor = (first, second) => {
            if (!first || !second) return null;
            const firstAncestors = new Set();
            let node = first;
            while (node) {
              firstAncestors.add(node);
              node = node.parentElement;
            }
            node = second;
            while (node) {
              if (firstAncestors.has(node)) return node;
              node = node.parentElement;
            }
            return null;
          };
          const domDistance = (first, second) => {
            if (!first || !second) return Number.MAX_SAFE_INTEGER;
            const firstAncestors = new Map();
            let node = first;
            let distance = 0;
            while (node) {
              firstAncestors.set(node, distance);
              node = node.parentElement;
              distance += 1;
            }
            node = second;
            distance = 0;
            while (node) {
              if (firstAncestors.has(node)) return distance + firstAncestors.get(node);
              node = node.parentElement;
              distance += 1;
            }
            return Number.MAX_SAFE_INTEGER;
          };
          const associatedLabels = labelCandidates.map(el => ({
            el,
            calendarWrapper: lowestCommonAncestor(el, verifiedCalendarRoot),
          })).filter(item => item.calendarWrapper &&
            item.calendarWrapper !== document.body &&
            item.calendarWrapper !== document.documentElement);
          associatedLabels.sort((a, b) => {
            const aDistance = domDistance(a.el, verifiedCalendarRoot);
            const bDistance = domDistance(b.el, verifiedCalendarRoot);
            if (aDistance !== bDistance) return aDistance - bDistance;
            const aChildren = a.el.querySelectorAll('*').length;
            const bChildren = b.el.querySelectorAll('*').length;
            if (aChildren !== bChildren) return aChildren - bChildren;
            const aBox = a.el.getBoundingClientRect();
            const bBox = b.el.getBoundingClientRect();
            return aBox.width * aBox.height - bBox.width * bBox.height;
          });
          const labelResult = associatedLabels[0] || null;
          const monthLabel = labelResult?.el || null;
          const calendarWrapper = labelResult?.calendarWrapper || null;
          if (!monthLabel) {
            return {
              count: 0,
              reason: verifiedCalendarRoot
                ? `No visible ${fullMonth} ${year} label shared a calendar wrapper with the verified grid`
                : 'No verified calendar grid root was available for month-header association',
              monthLabel: null,
              toolbar: null,
              candidates: [],
              selected: null,
            };
          }

          const isExplicitClickable = el => el.matches(explicitClickableSelector) &&
            !(el.hasAttribute('tabindex') && Number(el.getAttribute('tabindex')) < 0);
          const canonicalClickable = (node, region) => {
            const explicit = node.closest(explicitClickableSelector);
            if (explicit && region.contains(explicit) && isExplicitClickable(explicit)) {
              return explicit;
            }
            if (getComputedStyle(node).cursor === 'pointer') {
              let pointerControl = node;
              let parent = pointerControl.parentElement;
              while (
                parent && parent !== region && region.contains(parent) &&
                getComputedStyle(parent).cursor === 'pointer' &&
                !parent.contains(monthLabel)
              ) {
                pointerControl = parent;
                parent = parent.parentElement;
              }
              return pointerControl;
            }
            if (node.matches(iconSelector)) {
              const iconSignature = clean(
                `${node.textContent || ''} ${attributeText(node)}`
              ).toLowerCase();
              if (
                /next|prev|right|left|chevron|arrow|angle|caret|calendar|date_range|[<>›‹❯❮→←]/
                  .test(iconSignature)
              ) {
                return node;
              }
            }
            return null;
          };
          const semanticNodes = el => [
            el,
            ...el.querySelectorAll(
              'svg, use, i, mat-icon, span, [data-icon], [aria-label], [title]'
            ),
          ];
          const semanticText = el => clean(semanticNodes(el).map(node =>
            `${node.textContent || ''} ${attributeText(node)}`
          ).join(' ').replace(/([a-z0-9])([A-Z])/g, '$1 $2')).toLowerCase();
          const glyphPattern = {
            next: /^(?:>|›|❯|❱|»|→)$/,
            previous: /^(?:<|‹|❮|❰|«|←)$/,
          };
          const semanticPattern = {
            next: /\\bnext(?:\\s+month)?\\b|\\bnavigate[ _-]*next\\b|\\bkeyboard[ _-]*arrow[ _-]*right\\b|\\bchevron[ _-]*right\\b|\\barrow[ _-]*(?:right|forward)\\b|\\bangle[ _-]*right\\b|\\bcaret[ _-]*right\\b|\\bright[ _-]*nav\\b|\\bnav[ _-]*right\\b|\\bicon[ _-]*right\\b/,
            previous: /\\bprev(?:ious)?(?:\\s+month)?\\b|\\bnavigate[ _-]*(?:previous|before)\\b|\\bkeyboard[ _-]*arrow[ _-]*left\\b|\\bchevron[ _-]*left\\b|\\barrow[ _-]*(?:left|back)\\b|\\bangle[ _-]*left\\b|\\bcaret[ _-]*left\\b|\\bleft[ _-]*nav\\b|\\bnav[ _-]*left\\b|\\bicon[ _-]*left\\b/,
          };
          const calendarPattern = /\\bcalendar\\b|calendar[ _-]*(?:today|month)|calendar[ _-]*view[ _-]*month|date[ _-]*range|\\btoday\\b|\\bevent\\b/;
          const scoreDirection = (el, wantedDirection) => {
            const signature = semanticText(el);
            const text = clean(el.textContent);
            const ariaAndTitle = clean(
              `${el.getAttribute('aria-label') || ''} ${el.getAttribute('title') || ''}`
            ).toLowerCase();
            const classAndId = clean(
              `${el.id || ''} ${el.getAttribute('class') || ''}`
            ).toLowerCase();
            const opposite = wantedDirection === 'next' ? 'previous' : 'next';
            const reasons = [];
            let score = 0;
            const exactAccessible = wantedDirection === 'next'
              ? /\\bnext\\s+month\\b/.test(ariaAndTitle)
              : /\\bprev(?:ious)?\\s+month\\b/.test(ariaAndTitle);
            if (exactAccessible) {
              score = 1000;
              reasons.push('explicit accessible month-navigation label');
            }
            const fullCalendarClass = wantedDirection === 'next'
              ? /(?:^|\\s)fc-next(?:-button)?(?:\\s|$)/.test(classAndId)
              : /(?:^|\\s)fc-prev(?:-button)?(?:\\s|$)/.test(classAndId);
            if (fullCalendarClass && score < 950) {
              score = 950;
              reasons.push('FullCalendar navigation class');
            }
            if (glyphPattern[wantedDirection].test(text) && score < 850) {
              score = 850;
              reasons.push('directional chevron/arrow glyph');
            }
            if (semanticPattern[wantedDirection].test(signature) && score < 800) {
              score = 800;
              reasons.push('directional element or descendant semantics');
            }
            const hasOppositeSemantics = semanticPattern[opposite].test(signature) ||
              glyphPattern[opposite].test(text);
            if (hasOppositeSemantics && score < 900) {
              return { score: 0, reasons: ['opposite-direction semantics'] };
            }
            if (calendarPattern.test(signature) && score < 700) {
              return { score: 0, reasons: ['calendar control, not a direction control'] };
            }
            if (score && el.matches('button, a, [role="button"]')) score += 20;
            return { score, reasons };
          };
          const iconDetails = el => {
            const icons = [];
            const nodes = [
              ...(el.matches(iconSelector) ? [el] : []),
              ...el.querySelectorAll(iconSelector),
            ];
            for (const node of [...new Set(nodes)].filter(visible)) {
              icons.push({
                tag: node.tagName.toLowerCase(),
                text: clean(node.textContent),
                id: node.id || '',
                className: node.getAttribute('class') || '',
                title: node.getAttribute('title') || '',
                ariaLabel: node.getAttribute('aria-label') || '',
                href: node.getAttribute('href') || node.getAttribute('xlink:href') || '',
                dataAttributes: dataAttributes(node),
                outerHTML: node.outerHTML,
              });
            }
            return icons;
          };
          const clickablesWithin = region => {
            const all = [region, ...region.querySelectorAll('*')];
            const raw = [];
            for (const node of all) {
              if (!visible(node)) continue;
              if (isExplicitClickable(node) || getComputedStyle(node).cursor === 'pointer') {
                const canonical = canonicalClickable(node, region);
                if (canonical) raw.push(canonical);
              }
              if (node.matches(iconSelector)) {
                const canonical = canonicalClickable(node, region);
                if (canonical) raw.push(canonical);
              }
            }
            const unique = [...new Set(raw)].filter(visible)
              .filter(el => !el.closest(
                '[data-date], .fc-event, .fc-event-main, [data-event-id], [role="gridcell"]'
              ))
              .filter(el => el !== monthLabel && !el.contains(monthLabel));
            const eligible = unique.filter(enabled);
            const leaves = eligible.filter(el => !eligible.some(other =>
              other !== el && el.contains(other)));
            leaves.sort((a, b) => a === b ? 0 :
              (a.compareDocumentPosition(b) & Node.DOCUMENT_POSITION_FOLLOWING ? -1 : 1));
            unique.sort((a, b) => a === b ? 0 :
              (a.compareDocumentPosition(b) & Node.DOCUMENT_POSITION_FOLLOWING ? -1 : 1));
            return { diagnosticElements: unique, eligibleElements: leaves };
          };
          const describeCandidate = el => {
            const target = scoreDirection(el, direction);
            const oppositeDirection = direction === 'next' ? 'previous' : 'next';
            const opposite = scoreDirection(el, oppositeDirection);
            return {
              el,
              tag: el.tagName.toLowerCase(),
              text: clean(el.textContent),
              title: el.getAttribute('title') || '',
              ariaLabel: el.getAttribute('aria-label') || '',
              className: el.getAttribute('class') || '',
              id: el.id || '',
              href: el.getAttribute('href') || el.getAttribute('xlink:href') ||
                (typeof el.href === 'string' ? el.href : ''),
              role: el.getAttribute('role') || '',
              enabled: enabled(el),
              dataAttributes: dataAttributes(el),
              boundingBox: boundingBox(el),
              outerHTML: el.outerHTML,
              icons: iconDetails(el),
              semanticText: semanticText(el).slice(0, 1200),
              score: target.score,
              reasons: target.reasons,
              oppositeScore: opposite.score,
            };
          };
          const structuralChoice = items => {
            const calendarIndexes = items.map((item, index) =>
              calendarPattern.test(item.semanticText) ? index : -1
            ).filter(index => index >= 0);
            const choices = [];
            for (const calendarIndex of calendarIndexes) {
              const calendarControl = items[calendarIndex];
              const leftControl = items[calendarIndex - 1];
              const rightControl = items[calendarIndex + 1];
              if (!leftControl || !rightControl) continue;
              const sameParent = calendarControl.el.parentElement;
              if (
                leftControl.el.parentElement !== sameParent ||
                rightControl.el.parentElement !== sameParent
              ) continue;
              const leftLooksLikeIcon = leftControl.icons.length ||
                glyphPattern.previous.test(leftControl.text);
              const rightLooksLikeIcon = rightControl.icons.length ||
                glyphPattern.next.test(rightControl.text);
              if (!leftLooksLikeIcon || !rightLooksLikeIcon) continue;
              const candidate = direction === 'next' ? rightControl : leftControl;
              if (candidate.score || candidate.oppositeScore) continue;
              choices.push(candidate);
            }
            return choices.length === 1 ? choices[0] : null;
          };

          const ancestors = [];
          let ancestor = monthLabel.parentElement;
          while (ancestor && ancestor !== document.body && ancestors.length < 14) {
            if (visible(ancestor)) ancestors.push(ancestor);
            if (ancestor === calendarWrapper) break;
            ancestor = ancestor.parentElement;
          }
          const analyzedRegions = ancestors.map(region => {
            const clickables = clickablesWithin(region);
            const items = clickables.eligibleElements.map(describeCandidate);
            const diagnosticItems = clickables.diagnosticElements.map(describeCandidate);
            return {
              region,
              items,
              diagnosticItems,
              hasTarget: items.some(item => item.score > 0),
              hasOpposite: items.some(item => item.oppositeScore > 0),
              structural: structuralChoice(items),
              semanticRegion: /toolbar|header|month|calendar|navigation|nav/i.test(
                `${region.id || ''} ${region.getAttribute('class') || ''} ${region.getAttribute('role') || ''}`
              ),
            };
          });
          const regionResult = analyzedRegions.find(item =>
            item.hasTarget && item.hasOpposite
          ) || analyzedRegions.find(item =>
            item.hasTarget && item.semanticRegion
          ) || analyzedRegions.find(item => item.hasTarget) ||
            analyzedRegions.find(item => item.structural) ||
            analyzedRegions.find(item =>
              item.semanticRegion && item.diagnosticItems.length
            ) || analyzedRegions.find(item =>
              item.diagnosticItems.length >= 2
            ) || null;

          if (!regionResult) {
            return {
              count: 0,
              reason: 'No clickable elements were found near the verified month label',
              monthLabel: {
                tag: monthLabel.tagName.toLowerCase(),
                text: clean(monthLabel.textContent),
                boundingBox: boundingBox(monthLabel),
                outerHTML: monthLabel.outerHTML.slice(0, 2000),
              },
              toolbar: null,
              candidates: [],
              selected: null,
            };
          }

          const scored = regionResult.items.filter(item => item.score > 0)
            .sort((a, b) => b.score - a.score);
          let selected = null;
          let best = [];
          if (scored.length) {
            const bestScore = scored[0].score;
            best = scored.filter(item => item.score === bestScore);
            if (best.length === 1) selected = best[0];
          } else if (regionResult.structural) {
            selected = regionResult.structural;
            selected.score = 400;
            selected.reasons = [
              direction === 'next'
                ? 'right icon in a verified left/calendar/right toolbar triplet'
                : 'left icon in a verified left/calendar/right toolbar triplet',
            ];
            best = [selected];
          }
          if (selected) {
            selected.el.setAttribute('data-us-event-month-nav-target', 'true');
          }
          const serialize = item => {
            if (!item) return null;
            const { el, ...values } = item;
            return values;
          };
          return {
            count: selected ? 1 : best.length,
            reason: selected
              ? 'Unique verified month-navigation control found'
              : (best.length > 1
                ? 'Multiple month-navigation controls tied for the best score'
                : 'No verified directional control found in the month toolbar'),
            monthLabel: {
              tag: monthLabel.tagName.toLowerCase(),
              text: clean(monthLabel.textContent),
              boundingBox: boundingBox(monthLabel),
              outerHTML: monthLabel.outerHTML.slice(0, 2000),
            },
            toolbar: {
              tag: regionResult.region.tagName.toLowerCase(),
              id: regionResult.region.id || '',
              className: regionResult.region.getAttribute('class') || '',
              role: regionResult.region.getAttribute('role') || '',
              boundingBox: boundingBox(regionResult.region),
              outerHTML: regionResult.region.outerHTML.slice(0, 5000),
            },
            candidates: regionResult.diagnosticItems.map(serialize),
            selected: serialize(selected),
          };
        }
        """,
        {"direction": direction, "currentKey": current_key},
    )


def print_month_navigation_diagnostics(marked: dict) -> None:
    print("Month-label element:", marked.get("monthLabel"))
    toolbar = marked.get("toolbar")
    if toolbar:
        toolbar_summary = {
            key: toolbar.get(key, "")
            for key in ("tag", "id", "className", "role", "boundingBox")
        }
        print("Verified month-header region:", toolbar_summary)
    else:
        print("Verified month-header region: <not found>")
    print("Visible clickable elements in the month-header region:")
    candidates = marked.get("candidates", [])
    if not candidates:
        print("  (none)")
    for index, candidate in enumerate(candidates, start=1):
        print(f"Candidate {index}:")
        print(f"  tag: {candidate.get('tag', '')}")
        print(f"  text: {candidate.get('text', '')}")
        print(f"  title: {candidate.get('title', '')}")
        print(f"  aria-label: {candidate.get('ariaLabel', '')}")
        print(f"  class: {candidate.get('className', '')}")
        print(f"  id: {candidate.get('id', '')}")
        print(f"  href: {candidate.get('href', '')}")
        print(f"  enabled: {candidate.get('enabled', False)}")
        print(f"  bounding box: {candidate.get('boundingBox')}")
        print(f"  data attributes: {candidate.get('dataAttributes', {})}")
        print(f"  navigation score: {candidate.get('score', 0)}")
        print(f"  reasons: {candidate.get('reasons', [])}")
        print(f"  icon descendants: {candidate.get('icons', [])}")
        print(f"  outerHTML: {candidate.get('outerHTML', '')}")


def navigate_to_calendar_month(page, config: MonthConfig) -> None:
    for _ in range(36):
        state = read_calendar_month_state(page)
        current_key = state.get("monthKey", "")
        print(
            f"Calendar month detected: title={state.get('title', '')!r}, "
            f"month={current_key!r}, date-counts={state.get('monthCounts', {})}"
        )
        if current_key == config.prefix:
            print(f"Reached calendar month: {config.label}")
            return
        if not current_key:
            raise RuntimeError(
                f"Could not determine the displayed calendar month while navigating to {config.label}."
            )

        direction = "next" if current_key < config.prefix else "previous"
        before_key = current_key
        marked = mark_month_navigation_control(page, direction, before_key)
        print(f"Month-navigation detection: {marked.get('reason', '')}")
        print_month_navigation_diagnostics(marked)
        if marked.get("count") != 1 or not marked.get("selected"):
            raise RuntimeError(
                f"Expected one verified {direction}-month control, found "
                f"{marked.get('count', 0)}. {marked.get('reason', '')}"
            )

        before_year, before_month = (int(part) for part in before_key.split("-"))
        before_ordinal = before_year * 12 + before_month - 1
        expected_ordinal = before_ordinal + (1 if direction == "next" else -1)
        expected_key = f"{expected_ordinal // 12:04d}-{expected_ordinal % 12 + 1:02d}"

        selected = marked["selected"]
        print(f"Current detected month: {before_key}")
        print(f"About to click {direction.upper()} MONTH control")
        print(f"{direction.title()}-month element:")
        print(selected.get("outerHTML", ""))
        locator = page.locator('[data-us-event-month-nav-target="true"]')
        try:
            locator.click(timeout=10000)
        except PlaywrightError as normal_click_error:
            print(f"Normal Playwright click failed: {normal_click_error}")
            post_error_months = []
            before_month_reads = 0
            normal_click_changed_month = False
            for _ in range(20):
                page.wait_for_timeout(250)
                try:
                    post_error_state = read_calendar_month_state(page)
                except PlaywrightError:
                    post_error_months.append("<DOM read error>")
                    continue
                post_error_key = post_error_state.get("monthKey", "")
                post_error_months.append(post_error_key or "<undetected>")
                if post_error_key == expected_key:
                    normal_click_changed_month = True
                    break
                if post_error_key == before_key:
                    before_month_reads += 1
                elif post_error_key:
                    raise RuntimeError(
                        f"The normal {direction}-month click reported an error and "
                        f"the calendar moved unexpectedly to {post_error_key}; "
                        "refusing to click again."
                    ) from normal_click_error
            if normal_click_changed_month:
                print(
                    "The normal click already changed the month; skipping the "
                    "forced retry."
                )
            elif before_month_reads >= 3 and locator.count() == 1:
                print(
                    "Calendar remained on the previous month. Retrying the same "
                    "verified month-navigation element with force=True."
                )
                try:
                    locator.click(timeout=10000, force=True)
                except PlaywrightError as force_click_error:
                    raise RuntimeError(
                        "Both normal and forced clicks failed on the same verified "
                        f"{direction}-month control: {force_click_error}"
                    ) from force_click_error
            else:
                raise RuntimeError(
                    "Normal month-navigation click failed, and the page state was "
                    "not safe for a second click. Observed states: "
                    f"{post_error_months}."
                ) from normal_click_error

        changed = False
        consecutive_expected_reads = 0
        observed_months = []
        for _ in range(20):
            page.wait_for_timeout(250)
            try:
                new_state = read_calendar_month_state(page)
            except PlaywrightError as transient_error:
                observed_months.append(f"DOM read error: {transient_error}")
                consecutive_expected_reads = 0
                continue
            new_key = new_state.get("monthKey", "")
            observed_months.append(new_key or "<undetected>")
            if new_key == expected_key:
                consecutive_expected_reads += 1
                if consecutive_expected_reads >= 2:
                    changed = True
                    print(f"Previous month: {before_key}")
                    print(f"New month detected: {new_key}")
                    print(f"{direction.title()}-month navigation PASSED")
                    break
            else:
                consecutive_expected_reads = 0
        if not changed:
            raise RuntimeError(
                f"The verified {direction}-month control did not move from {before_key} "
                f"to {expected_key}. Observed month states: {observed_months}. "
                "Stopping before scraping the requested month."
            )

    raise RuntimeError(f"Exceeded safe month-navigation limit while seeking {config.label}.")


def wait_for_month_calendar(page, config: MonthConfig) -> None:
    try:
        page.wait_for_function(
            """
        target => {
          const root = document.querySelector('[data-us-event-calendar-root="true"]') || document;
          const elements = [root, ...root.querySelectorAll('*')];
          const calendarDayDates = new Set(elements.flatMap(el => [...el.attributes]
            .filter(attr => attr.name.toLowerCase() === 'data-date' && /^20\\d{2}-\\d{2}-\\d{2}$/.test(attr.value))
            .map(attr => attr.value)));
          for (let day = 1; day <= target.days; day += 1) {
            const expected = `${target.prefix}-${String(day).padStart(2, '0')}`;
            if (!calendarDayDates.has(expected)) return false;
          }
          return true;
        }
        """,
            arg={"prefix": config.prefix, "days": config.days_in_month},
            timeout=30000,
        )
    except PlaywrightTimeoutError as exc:
        fallback_state = read_calendar_month_state(page)
        fallback_supported, fallback_reason = sparse_event_fallback_support(
            fallback_state, config
        )
        if not fallback_supported:
            raise RuntimeError(
                f"The {config.label} calendar did not render every target day cell, "
                f"and the safe sparse-grid fallback was unavailable: {fallback_reason}."
            ) from exc
        print(
            f"{config.label} uses a sparse calendar DOM; target day cells and "
            "FullCalendar event metadata agree."
        )

    # FullCalendar renders asynchronously. Require the semantic event/date
    # counts to remain unchanged across three checks before scraping.
    previous_counts = None
    stable_checks = 0
    for _ in range(20):
        counts = page.evaluate(
            """
            target => {
              const root = document.querySelector('[data-us-event-calendar-root="true"]') || document;
              const elements = [root, ...root.querySelectorAll('*')];
              return {
                events: elements.filter(el => Boolean(el.fcSeg)).length,
                targetDates: elements.filter(el => [...el.attributes]
                  .some(attr => /date|start/i.test(attr.name) && attr.value.includes(target.prefix))).length,
                gigListings: elements.filter(el =>
                  /^(?:1[0-2]|0?[1-9])(?::[0-5]\\d)?\\s*[ap](?:m)?\\s*[a-z]{2}\\s*\\|\\|/i
                    .test((el.textContent || '').replace(/\\s+/g, ' ').trim())).length,
                moreLinks: elements.filter(el => /^\\+\\d+\\s+more$/i
                  .test((el.textContent || '').replace(/\\s+/g, ' ').trim())).length,
              };
            }
            """,
            {"prefix": config.prefix},
        )
        if counts == previous_counts:
            stable_checks += 1
            if stable_checks >= 3:
                print(
                    f"{config.label} calendar DOM stabilized: "
                    f"{counts['events']} FullCalendar event nodes, "
                    f"{counts['targetDates']} target-month date-bearing nodes, "
                    f"{counts['gigListings']} semantic gig listings."
                )
                return
        else:
            previous_counts = counts
            stable_checks = 0
        page.wait_for_timeout(500)

    raise RuntimeError(f"The {config.label} calendar DOM did not stabilize before scraping.")


def inspect_calendar_dom(page, config: MonthConfig) -> dict:
    month_state = validate_calendar_month_state(
        read_calendar_month_state(page), config
    )
    report = page.evaluate(
        """
        target => {
          const verifiedCalendarRoot = document.querySelector(
            '[data-us-event-calendar-root="true"]'
          );
          const inspectionScope = verifiedCalendarRoot || document;
          const elements = [inspectionScope, ...inspectionScope.querySelectorAll('*')];
          const clean = value => String(value || '').replace(/\\s+/g, ' ').trim();
          const targetDateElements = elements.filter(el => [...el.attributes]
            .some(attr => /date|start/i.test(attr.name) && attr.value.includes(target.prefix)));
          const eventElements = elements.filter(el => Boolean(el.fcSeg) &&
            !String(el.getAttribute('class') || '').toLowerCase().includes('event-mirror'));
          const moreElements = elements.filter(el => {
            const text = clean(el.textContent);
            const classes = String(el.getAttribute('class') || '').toLowerCase();
            return /^\\+\\d+\\s+more$/i.test(text) || classes.includes('more-link');
          });

          const semanticRoots = elements.filter(el => {
            const semantic = `${el.tagName} ${el.id || ''} ${el.getAttribute('class') || ''}`.toLowerCase();
            return semantic.includes('calendar') || semantic.includes('full-calendar') ||
              semantic.includes('fc-view') || semantic.includes('fc-daygrid') || semantic.includes('rbc-calendar');
          }).map(el => ({
            el,
            dateCount: targetDateElements.filter(dateEl => el.contains(dateEl)).length,
          })).filter(item => item.dateCount > 0)
            .sort((a, b) => b.dateCount - a.dateCount || a.el.querySelectorAll('*').length - b.el.querySelectorAll('*').length);

          const root = semanticRoots[0]?.el || targetDateElements[0]?.parentElement || null;
          const firstDateAncestors = [];
          let ancestor = targetDateElements[0] || null;
          while (ancestor) {
            firstDateAncestors.push(ancestor);
            ancestor = ancestor.parentElement;
          }
          const calendarRoot = verifiedCalendarRoot || firstDateAncestors.find(el =>
            el.tagName.toLowerCase().includes('calendar')
          ) || firstDateAncestors.find(el =>
            /full-calendar|calendar-container|calendar-wrapper/i.test(
              `${el.id || ''} ${el.getAttribute('class') || ''}`
            )
          ) || root;
          document.querySelectorAll('[data-us-event-calendar-root]').forEach(el =>
            el.removeAttribute('data-us-event-calendar-root'));
          if (calendarRoot) calendarRoot.setAttribute('data-us-event-calendar-root', 'true');
          const classTokens = [...new Set(elements.flatMap(el =>
            String(el.getAttribute('class') || '').split(/\\s+/)
              .filter(token => /calendar|^fc-|^rbc-/i.test(token))
          ))].slice(0, 100);

          const attributeInventory = {};
          for (const el of elements) {
            for (const attr of el.attributes) {
              if (/date|start|event/i.test(attr.name)) {
                attributeInventory[attr.name] = (attributeInventory[attr.name] || 0) + 1;
              }
            }
          }

          const samples = eventElements.slice(0, 20).map(el => {
            const seg = el.fcSeg;
            const def = seg?.eventRange?.def;
            const instance = seg?.eventRange?.instance;
            const range = instance?.range || seg?.eventRange?.range;
            const start = range?.start;
            return {
              tag: el.tagName.toLowerCase(),
              id: el.id || '',
              className: el.getAttribute('class') || '',
              eventId: instance?.instanceId || def?.publicId || def?.defId || '',
              title: def?.title || '',
              start: start && typeof start.toISOString === 'function' ? start.toISOString() : clean(start),
              dayCellDate: el.closest('[data-date]')?.getAttribute('data-date') || '',
              ownMetadataKeys: Object.getOwnPropertyNames(el)
                .filter(key => /seg|event|calendar/i.test(key)),
              outerHTML: el.outerHTML.slice(0, 1600),
            };
          });

          return {
            root: calendarRoot ? {
              tag: calendarRoot.tagName.toLowerCase(),
              id: calendarRoot.id || '',
              className: calendarRoot.getAttribute('class') || '',
              outerHTML: calendarRoot.outerHTML.slice(0, 3000),
            } : null,
            monthTitle: clean(elements.find(el =>
              String(el.getAttribute('class') || '').toLowerCase().includes('toolbar-title')
            )?.textContent),
            distinctTargetDayDates: [...new Set(elements.flatMap(el => [...el.attributes]
              .filter(attr => attr.name.toLowerCase() === 'data-date' && attr.value.startsWith(target.prefix + '-'))
              .map(attr => attr.value)))].sort(),
            libraryClassTokens: classTokens,
            attributeInventory,
            targetDateElementCount: targetDateElements.length,
            fullCalendarEventNodeCount: eventElements.length,
            uniqueInstanceCount: new Set(eventElements.map(el =>
              el.fcSeg?.eventRange?.instance?.instanceId || ''
            ).filter(Boolean)).size,
            moreLinkCount: moreElements.length,
            moreLinks: moreElements.map(el => ({
              text: clean(el.textContent),
              date: el.closest('[data-date]')?.getAttribute('data-date') || '',
              className: el.getAttribute('class') || '',
              ownMetadataKeys: Object.getOwnPropertyNames(el)
                .filter(key => /seg|event|more|calendar/i.test(key)),
              outerHTML: el.outerHTML.slice(0, 1000),
            })),
            eventSamples: samples,
          };
        }
        """,
        {"prefix": config.prefix},
    )

    report["monthTitle"] = month_state.get("title", "")
    report["monthValidation"] = month_state

    print("CALENDAR DOM INSPECTION")
    root_summary = {
        key: report["root"].get(key, "")
        for key in ("tag", "id", "className")
    } if report["root"] else None
    print("Calendar root:", root_summary)
    print(
        f"Calendar DOM counts: {len(month_state.get('dayCellDates', []))} distinct "
        f"day cells, {report['fullCalendarEventNodeCount']} event nodes, "
        f"{report['uniqueInstanceCount']} unique event instances, "
        f"{report['moreLinkCount']} overflow controls."
    )
    for more_link in report["moreLinks"]:
        print(
            "Overflow control:",
            {"text": more_link.get("text", ""), "date": more_link.get("date", "")},
        )
    for sample in report["eventSamples"][:5]:
        print(
            "Event metadata sample:",
            {
                "eventId": sample.get("eventId", ""),
                "title": sample.get("title", ""),
                "start": sample.get("start", ""),
                "dayCellDate": sample.get("dayCellDate", ""),
            },
        )
    return report


def extract_calendar_event_dom(page) -> dict:
    return page.evaluate(
        """
        () => {
          const calendarRoot = document.querySelector('[data-us-event-calendar-root="true"]');
          if (!calendarRoot) {
            throw new Error('The verified calendar root marker is missing.');
          }
          const scopedElements = [calendarRoot, ...calendarRoot.querySelectorAll('*')];
          const hasActiveOverflow = Boolean(document.querySelector('[data-us-event-overflow-target="true"]'));
          const overflowPopoverElements = hasActiveOverflow
            ? [...document.querySelectorAll('*')].filter(el => {
                const signature = `${el.getAttribute('class') || ''} ${el.getAttribute('role') || ''}`.toLowerCase();
                return /popover|dialog/.test(signature) && [...el.querySelectorAll('*')].some(child => Boolean(child.fcSeg));
              }).flatMap(popover => [popover, ...popover.querySelectorAll('*')])
            : [];
          const elements = [...new Set([...scopedElements, ...overflowPopoverElements])];
          const clean = value => String(value || '').replace(/\\s+/g, ' ').trim();
          const absoluteUrl = value => {
            if (!value) return '';
            try { return new URL(value, location.href).href; } catch (_) { return String(value); }
          };
          const isoValue = value => {
            if (!value) return '';
            try {
              if (typeof value.toISOString === 'function') return value.toISOString();
            } catch (_) {}
            return clean(value);
          };
          const dateFromString = value => {
            const match = clean(value).match(/\\b(20\\d{2}-\\d{2}-\\d{2})\\b/);
            return match ? match[1] : '';
          };
          const meaningfulColor = value => {
            const color = clean(value);
            if (!color || /transparent|rgba\\([^)]*,\\s*0(?:\\.0+)?\\)/i.test(color)) return '';
            return color;
          };
          const calendarColor = (el, ui = {}) => {
            const ownStyle = getComputedStyle(el);
            const descendants = [...el.querySelectorAll('*')];
            const dot = descendants.find(node => {
              const signature = `${node.getAttribute('class') || ''} ${node.getAttribute('style') || ''}`.toLowerCase();
              return /event-dot|background-color|border-color/.test(signature);
            });
            const dotStyle = dot ? getComputedStyle(dot) : null;
            return [
              ui.backgroundColor,
              ui.borderColor,
              el.style.backgroundColor,
              el.style.borderColor,
              ownStyle.backgroundColor,
              ownStyle.borderLeftColor,
              ownStyle.borderColor,
              dotStyle?.backgroundColor,
              dotStyle?.borderLeftColor,
              dotStyle?.borderColor,
            ].map(meaningfulColor).find(Boolean) || '';
          };
          const isLayoutVisible = el => {
            let node = el;
            while (node && node !== document.body) {
              const style = getComputedStyle(node);
              if (style.display === 'none' || style.visibility === 'hidden' ||
                  style.visibility === 'collapse' || Number(style.opacity || '1') === 0) {
                return false;
              }
              node = node.parentElement;
            }
            return el.getClientRects().length > 0;
          };
          const serializableProps = props => {
            if (!props || typeof props !== 'object') return {};
            const output = {};
            for (const [key, value] of Object.entries(props)) {
              if (value == null || ['string', 'number', 'boolean'].includes(typeof value)) {
                output[key] = value;
              }
            }
            return output;
          };
          const textBySemanticClass = (el, semantic) => {
            const node = [...el.querySelectorAll('*')].find(child =>
              String(child.getAttribute('class') || '').toLowerCase().includes(semantic)
            );
            return node ? clean(node.textContent) : '';
          };
          const eventUrl = (el, metadataUrl = '') => {
            if (metadataUrl) return absoluteUrl(metadataUrl);
            if (el.matches('a[href]')) return absoluteUrl(el.getAttribute('href'));
            const link = el.querySelector('a[href]') || el.closest('a[href]');
            return link ? absoluteUrl(link.getAttribute('href')) : '';
          };

          const fcEventElements = elements.filter(el => Boolean(el.fcSeg) &&
            el.fcSeg?.isStart !== false &&
            !String(el.getAttribute('class') || '').toLowerCase().includes('event-mirror'));
          const fcEvents = fcEventElements.map(el => {
            const seg = el.fcSeg;
            const eventRange = seg?.eventRange || {};
            const def = eventRange.def || {};
            const instance = eventRange.instance || {};
            const range = instance.range || eventRange.range || {};
            const startIso = isoValue(range.start);
            const metadataDate = dateFromString(startIso);
            const segmentStartIso = isoValue(eventRange.range?.start);
            const segmentDate = dateFromString(segmentStartIso);
            const dayCellDate = el.closest('[data-date]')?.getAttribute('data-date') || '';
            const timeText = textBySemanticClass(el, 'event-time');
            const titleText = textBySemanticClass(el, 'event-title') || clean(def.title);
            const rawText = clean([timeText, titleText].filter(Boolean).join(' ')) || clean(el.textContent);
            const stablePublicId = clean(
              def.publicId || def.extendedProps?.eventId || def.extendedProps?.event_id ||
              def.extendedProps?.gigId || def.extendedProps?.gig_id ||
              def.extendedProps?.shiftId || def.extendedProps?.shift_id || ''
            );
            return {
              source: 'FullCalendar fcSeg',
              eventId: instance.instanceId || stablePublicId || def.defId || '',
              publicId: stablePublicId,
              date: dayCellDate || segmentDate || metadataDate,
              metadataDate,
              segmentDate,
              dayCellDate,
              start: startIso,
              displayedTime: timeText,
              title: clean(def.title) || titleText,
              rawText,
              url: eventUrl(el, def.url || ''),
              color: calendarColor(el, eventRange.ui || {}),
              extendedProps: serializableProps(def.extendedProps),
              isVisible: isLayoutVisible(el),
              outerHTML: el.outerHTML.slice(0, 2400),
            };
          });

          const listingPattern = /^(?:1[0-2]|0?[1-9])(?::[0-5]\\d)?\\s*[ap](?:m)?\\s*[a-z]{2}\\s*\\|\\|/i;
          const listingText = el => clean(el.textContent);
          const strictMatches = elements.filter(el => listingPattern.test(listingText(el)));
          const smallestMatches = strictMatches.filter(el =>
            ![...el.children].some(child => listingPattern.test(listingText(child)))
          );
          const countListingMarkers = el => (listingText(el).match(/[a-z]{2}\\s*\\|\\|/gi) || []).length;
          const isEventWrapper = el => {
            const classes = String(el.getAttribute('class') || '').toLowerCase();
            const role = String(el.getAttribute('role') || '').toLowerCase();
            const hasEventData = [...el.attributes].some(attr => /data-(?:event|start)/i.test(attr.name));
            return el.matches('a, button') || role === 'link' || hasEventData ||
              classes.split(/\\s+/).some(token => /(?:^|[-_])(?:event|gig|shift)(?:$|[-_])/.test(token));
          };
          const closestEventWrapper = leaf => {
            let node = leaf;
            let fallback = leaf;
            while (node && node !== document.body) {
              if (countListingMarkers(node) > 1) break;
              fallback = node;
              if (isEventWrapper(node)) return node;
              node = node.parentElement;
            }
            return fallback;
          };
          const dateEvidence = root => {
            let node = root;
            while (node && node !== document.body) {
              const preferred = [...node.attributes].filter(attr => /(?:^|[-_:])(date|start)(?:$|[-_:])/i.test(attr.name));
              const other = [...node.attributes].filter(attr => /date|start|event/i.test(attr.name));
              for (const attr of [...preferred, ...other]) {
                const found = dateFromString(attr.value);
                if (found) return { date: found, source: `${node.tagName.toLowerCase()}[${attr.name}]` };
              }
              node = node.parentElement;
            }
            const datedTime = root.querySelector('time[datetime]');
            if (datedTime) {
              const found = dateFromString(datedTime.getAttribute('datetime'));
              if (found) return { date: found, source: 'time[datetime]' };
            }
            return { date: '', source: '' };
          };
          const startEvidence = root => {
            let node = root;
            while (node && node !== document.body) {
              const attr = [...node.attributes].find(item => /(?:^|[-_:])start(?:$|[-_:])/i.test(item.name));
              if (attr) return attr.value;
              node = node.parentElement;
            }
            return root.querySelector('time[datetime]')?.getAttribute('datetime') || '';
          };

          const fallbackEvents = [];
          const unresolved = [];
          for (const leaf of smallestMatches) {
            if (fcEventElements.some(eventEl => eventEl.contains(leaf))) continue;
            const wrapper = closestEventWrapper(leaf);
            const evidence = dateEvidence(wrapper);
            const rawText = listingText(wrapper);
            const stableDomId = wrapper.getAttribute('data-event-id') ||
              wrapper.getAttribute('data-gig-id') || wrapper.getAttribute('data-shift-id') || '';
            const entry = {
              source: 'semantic DOM fallback',
              eventId: stableDomId,
              publicId: stableDomId,
              date: evidence.date,
              metadataDate: '',
              dayCellDate: wrapper.closest('[data-date]')?.getAttribute('data-date') || '',
              dateSource: evidence.source,
              start: startEvidence(wrapper),
              displayedTime: (rawText.match(/^(?:1[0-2]|0?[1-9])(?::[0-5]\\d)?\\s*[ap](?:m)?/i) || [''])[0],
              title: '',
              rawText,
              url: eventUrl(wrapper),
              color: calendarColor(wrapper),
              extendedProps: {},
              isVisible: isLayoutVisible(wrapper),
              outerHTML: wrapper.outerHTML.slice(0, 2400),
            };
            if (entry.date) fallbackEvents.push(entry);
            else unresolved.push(entry);
          }

          return {
            events: [...fcEvents, ...fallbackEvents],
            unresolved,
            fcSegEventCount: fcEvents.length,
            fallbackEventCount: fallbackEvents.length,
          };
        }
        """
    )


def discover_overflow_links(page) -> list[dict]:
    return page.evaluate(
        """
        () => {
          const clean = value => String(value || '').replace(/\\s+/g, ' ').trim();
          const calendarRoot = document.querySelector('[data-us-event-calendar-root="true"]');
          if (!calendarRoot) throw new Error('Verified calendar root is missing.');
          const links = [...calendarRoot.querySelectorAll('a, button, [role="button"]')]
            .filter(el => /^\\+\\d+\\s+more$/i.test(clean(el.textContent)));
          return links.map((el, index) => {
            el.setAttribute('data-us-event-overflow-index', String(index));
            const text = clean(el.textContent);
            return {
              index,
              text,
              hiddenCount: Number((text.match(/\\d+/) || ['0'])[0]),
              date: el.closest('[data-date]')?.getAttribute('data-date') || '',
              outerHTML: el.outerHTML.slice(0, 1200),
            };
          });
        }
        """
    )


def mark_overflow_link_for_click(page, link: dict) -> int:
    return page.evaluate(
        """
        target => {
          const clean = value => String(value || '').replace(/\\s+/g, ' ').trim();
          const calendarRoot = document.querySelector('[data-us-event-calendar-root="true"]');
          if (!calendarRoot) return 0;
          document.querySelectorAll('[data-us-event-overflow-target]').forEach(el =>
            el.removeAttribute('data-us-event-overflow-target'));
          const matches = [...calendarRoot.querySelectorAll('a, button, [role="button"]')]
            .filter(el => clean(el.textContent) === target.text)
            .filter(el => (el.closest('[data-date]')?.getAttribute('data-date') || '') === target.date);
          if (matches.length === 1) {
            matches[0].setAttribute('data-us-event-overflow-target', 'true');
          }
          return matches.length;
        }
        """,
        {"date": link["date"], "text": link["text"]},
    )


def close_calendar_overflow_popover(page) -> None:
    page.keyboard.press("Escape")
    page.wait_for_timeout(250)
    close_count = page.evaluate(
        """
        () => {
          document.querySelectorAll('[data-us-event-popover-close]').forEach(el =>
            el.removeAttribute('data-us-event-popover-close'));
          const visible = el => {
            const style = getComputedStyle(el);
            const rect = el.getBoundingClientRect();
            return style.display !== 'none' && style.visibility !== 'hidden' && rect.width > 0 && rect.height > 0;
          };
          const popovers = [...document.querySelectorAll('*')].filter(el => {
            const signature = `${el.getAttribute('class') || ''} ${el.getAttribute('role') || ''}`.toLowerCase();
            return visible(el) && /popover|dialog/.test(signature) && el.querySelector('[class*="event" i]');
          });
          for (const popover of popovers) {
            const close = [...popover.querySelectorAll('button, [role="button"], [aria-label], [title], [class]')]
              .find(el => /close|dismiss|popover-close/.test(
                `${el.getAttribute('aria-label') || ''} ${el.getAttribute('title') || ''} ${el.getAttribute('class') || ''}`.toLowerCase()
              ));
            if (close) {
              close.setAttribute('data-us-event-popover-close', 'true');
              return 1;
            }
          }
          return 0;
        }
        """
    )
    if close_count:
        page.locator('[data-us-event-popover-close="true"]').click(timeout=5000)
        page.wait_for_timeout(250)


def collect_all_month_dom_events(page, config: MonthConfig) -> list[dict]:
    initial = extract_calendar_event_dom(page)
    if initial["unresolved"]:
        print("Unresolved gig-like DOM elements:")
        for item in initial["unresolved"]:
            print(item)
        raise RuntimeError(
            "At least one gig-like calendar element had no authoritative DOM date; refusing to create an incomplete spreadsheet."
        )

    def verify_event_semantics(extraction: dict, context: str) -> None:
        ambiguous = []
        for event in extraction["events"]:
            if not clean_text(event.get("date", "")).startswith(config.prefix + "-"):
                continue
            raw = clean_text(event.get("rawText", ""))
            time_match = GIG_TIME_PREFIX_RE.match(raw)
            remainder = raw[time_match.end():] if time_match else raw
            if not time_match or not STATE_PREFIX_RE.match(remainder):
                ambiguous.append(event)
        if ambiguous:
            print(f"Ambiguous non-gig FullCalendar events during {context}:")
            for event in ambiguous:
                print(event)
            raise RuntimeError(
                f"A FullCalendar event in {config.label} did not match the verified gig "
                "syntax (time + STATE||); refusing to mix non-gig entries into the export."
            )

    verify_event_semantics(initial, "initial calendar extraction")

    all_events = list(initial["events"])
    overflow_links = discover_overflow_links(page)
    initial_keys_by_date: dict[str, set[tuple[str, str, str]]] = {}
    initial_visible_keys_by_date: dict[str, set[tuple[str, str, str]]] = {}
    for event in all_events:
        key = (
            clean_text(event.get("eventId", "")),
            clean_text(event.get("start", "")),
            clean_text(event.get("rawText", "")).casefold(),
        )
        initial_keys_by_date.setdefault(event.get("date", ""), set()).add(key)
        if event.get("isVisible"):
            initial_visible_keys_by_date.setdefault(event.get("date", ""), set()).add(key)

    for link in overflow_links:
        if not link["date"]:
            print("Overflow link without an authoritative data-date:", link)
            raise RuntimeError(
                f"A calendar +more control had no authoritative day-cell date; refusing "
                f"to skip possible {config.label} gigs."
            )
        if not link["date"].startswith(config.prefix):
            continue
        visible_before = len(initial_visible_keys_by_date.get(link["date"], set()))
        initial_total = len(initial_keys_by_date.get(link["date"], set()))
        required_total = visible_before + link["hiddenCount"]
        if initial_total >= required_total:
            print(
                f"All events behind {link['text']} for {link['date']} are already present in the semantic DOM; no overflow click needed."
            )
            continue
        print(
            f"Opening read-only overflow control for {link['date']}: "
            f"{link['text']} (no gig element will be clicked)"
        )
        matched_links = mark_overflow_link_for_click(page, link)
        if matched_links != 1:
            raise RuntimeError(
                f"Could not uniquely relocate overflow control for {link['date']}; found {matched_links}."
            )
        locator = page.locator('[data-us-event-overflow-target="true"]')
        locator.click(timeout=10000)
        page.wait_for_timeout(500)
        expanded = extract_calendar_event_dom(page)
        if expanded["unresolved"]:
            print("Unresolved gig-like elements after opening overflow:")
            for item in expanded["unresolved"]:
                print(item)
            raise RuntimeError(
                f"Overflow events for {link['date']} lacked authoritative DOM dates."
            )
        verify_event_semantics(expanded, f"overflow expansion for {link['date']}")
        all_events.extend(expanded["events"])
        close_calendar_overflow_popover(page)

    # Validate that every +N link contributed at least N unique event instances
    # unless those instances were already present in the original DOM.
    combined_keys_by_date: dict[str, set[tuple[str, str, str]]] = {}
    for event in all_events:
        key = (
            clean_text(event.get("eventId", "")),
            clean_text(event.get("start", "")),
            clean_text(event.get("rawText", "")).casefold(),
        )
        combined_keys_by_date.setdefault(event.get("date", ""), set()).add(key)
    for link in overflow_links:
        if not link["date"]:
            continue
        if not link["date"].startswith(config.prefix):
            continue
        visible_before = len(initial_visible_keys_by_date.get(link["date"], set()))
        after = len(combined_keys_by_date.get(link["date"], set()))
        required_total = visible_before + link["hiddenCount"]
        if after < required_total:
            raise RuntimeError(
                f"The {link['text']} control for {link['date']} exposed only "
                f"{after} of at least {required_total} expected unique events; refusing an incomplete export."
            )

    print(
        f"{config.label} raw calendar extraction: "
        f"{len(all_events)} DOM renderings collected before deduplication."
    )
    return all_events


GIG_TIME_PREFIX_RE = re.compile(
    r"""
    ^\s*
    (?P<hour>1[0-2]|0?[1-9])
    (?:\s*:\s*(?P<minute>[0-5]\d))?
    \s*(?P<ampm>a(?:\.?\s*m\.?)?|p(?:\.?\s*m\.?)?)
    (?=\s|$|[|–—-]|[A-Za-z]{2}\s*\|\|)
    """,
    re.IGNORECASE | re.VERBOSE,
)
STATE_PREFIX_RE = re.compile(r"^\s*(?P<state>[A-Za-z]{2})\s*\|\|\s*")
CITY_SUFFIX_RE = re.compile(r"\s*\((?P<city>[^()\r\n]+)\)\s*$")
STORE_NUMBER_SUFFIX_RE = re.compile(
    r"""
    ^(?P<account>.+?\S)
    \s+
    (?:(?:store|location)\s*(?:no\.?|number)?\s*)?
    \#?\s*(?P<number>\d{1,10})\s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)


def parse_short_calendar_time(value: str) -> tuple[str, str]:
    match = GIG_TIME_PREFIX_RE.match(clean_text(value))
    if not match:
        return "", ""

    hour = int(match.group("hour"))
    minute = int(match.group("minute") or "0")
    ampm = re.sub(r"[^ap]", "", match.group("ampm").lower())
    hour_24 = hour % 12 + (12 if ampm.startswith("p") else 0)
    display_hour = hour_24 % 12 or 12
    display_suffix = "PM" if hour_24 >= 12 else "AM"
    return f"{display_hour}:{minute:02d} {display_suffix}", f"{hour_24:02d}:{minute:02d}"


def parse_dom_start_time(value: str) -> tuple[str, str]:
    value = clean_text(value)
    if not value:
        return "", ""

    iso_match = re.search(r"T(?P<hour>[01]\d|2[0-3]):(?P<minute>[0-5]\d)", value)
    if iso_match:
        hour_24 = int(iso_match.group("hour"))
        minute = int(iso_match.group("minute"))
        display_hour = hour_24 % 12 or 12
        suffix = "PM" if hour_24 >= 12 else "AM"
        return f"{display_hour}:{minute:02d} {suffix}", f"{hour_24:02d}:{minute:02d}"

    twelve_hour_match = re.search(
        r"(?P<hour>1[0-2]|0?[1-9]):(?P<minute>[0-5]\d)\s*(?P<ampm>AM|PM)\b",
        value,
        re.IGNORECASE,
    )
    if twelve_hour_match:
        return parse_short_calendar_time(
            f"{twelve_hour_match.group('hour')}:{twelve_hour_match.group('minute')}"
            f"{twelve_hour_match.group('ampm')[0]}"
        )
    return "", ""


def compact_calendar_time(start_time_sort: str) -> str:
    if not re.fullmatch(r"\d{2}:\d{2}", start_time_sort):
        return ""
    hour_24, minute = (int(part) for part in start_time_sort.split(":"))
    hour = hour_24 % 12 or 12
    suffix = "p" if hour_24 >= 12 else "a"
    return f"{hour}{':' + str(minute).zfill(2) if minute else ''}{suffix}"


def metadata_value(props: dict, *names: str) -> str:
    if not isinstance(props, dict):
        return ""
    lowered = {str(key).casefold(): value for key, value in props.items()}
    for name in names:
        value = lowered.get(name.casefold())
        if value is not None and not isinstance(value, (dict, list, tuple)):
            return clean_text(str(value))
    return ""


def parse_calendar_listing(event: dict, scraped_at: str, config: MonthConfig) -> dict:
    raw_listing = clean_text(event.get("rawText", ""))
    displayed_time = clean_text(event.get("displayedTime", ""))
    dom_start = clean_text(event.get("start", ""))

    start_time, start_time_sort = parse_short_calendar_time(displayed_time)
    if not start_time:
        start_time, start_time_sort = parse_short_calendar_time(raw_listing)
    dom_start_time, dom_start_sort = parse_dom_start_time(dom_start)
    if not start_time:
        start_time, start_time_sort = dom_start_time, dom_start_sort
    elif dom_start_sort and start_time_sort != dom_start_sort:
        print(
            "WARNING: visible time and FullCalendar start metadata disagree; "
            f"using visible listing time. raw={raw_listing!r} "
            f"visible={start_time_sort} metadata={dom_start_sort}"
        )

    if displayed_time and raw_listing and not GIG_TIME_PREFIX_RE.match(raw_listing):
        raw_listing = clean_text(f"{displayed_time} {raw_listing}")
    elif not raw_listing:
        title = clean_text(event.get("title", ""))
        if start_time_sort and title:
            raw_listing = clean_text(f"{compact_calendar_time(start_time_sort)} {title}")
        else:
            raw_listing = title

    remainder = raw_listing
    time_match = GIG_TIME_PREFIX_RE.match(remainder)
    if time_match:
        remainder = remainder[time_match.end():].strip()

    state = ""
    state_match = STATE_PREFIX_RE.match(remainder)
    if state_match:
        state = state_match.group("state").upper()
        remainder = remainder[state_match.end():].strip()

    city = ""
    city_match = CITY_SUFFIX_RE.search(remainder)
    if city_match:
        city = clean_text(city_match.group("city"))
        remainder = remainder[:city_match.start()].strip()

    store_number = ""
    store_account = remainder
    number_match = STORE_NUMBER_SUFFIX_RE.match(remainder)
    if number_match:
        store_account = clean_text(number_match.group("account"))
        store_number = number_match.group("number")

    props = event.get("extendedProps", {})
    if not state:
        candidate = metadata_value(props, "state", "storeState", "locationState")
        if re.fullmatch(r"[A-Za-z]{2}", candidate):
            state = candidate.upper()
    if not city:
        city = metadata_value(props, "city", "storeCity", "locationCity")
    if not store_number:
        store_number = metadata_value(
            props, "storeNumber", "store_number", "locationNumber", "location_number"
        )
    if not store_account:
        store_account = metadata_value(
            props, "store", "account", "storeName", "accountName", "locationName"
        )

    preferred_priorities = {
        "bellevue": 1,
        "kirkland": 2,
        "seattle": 3,
    }
    location_priority = preferred_priorities.get(city.casefold(), 9)
    preferred_area = "YES" if location_priority in {1, 2, 3} else "NO"

    event_date = clean_text(event.get("date", ""))
    try:
        parsed_date = date.fromisoformat(event_date)
    except ValueError as exc:
        raise RuntimeError(
            f"Calendar event has an invalid authoritative date {event_date!r}: {raw_listing!r}"
        ) from exc
    if parsed_date.strftime("%Y-%m") != config.prefix:
        raise RuntimeError(
            f"parse_calendar_listing received an event outside {config.label}: {event_date}"
        )

    return {
        "date": event_date,
        "day": parsed_date.strftime("%A"),
        "start_time": start_time,
        "start_time_sort": start_time_sort,
        "state": state,
        "store_account": store_account,
        "store_number": store_number,
        "city": city,
        "preferred_area": preferred_area,
        "location_priority": location_priority,
        "raw_listing": raw_listing,
        "calendar_color": clean_text(event.get("color", "")),
        "event_url": clean_text(event.get("url", "")),
        "scraped_at": scraped_at,
        "event_id": clean_text(event.get("eventId", "")),
        "public_id": clean_text(event.get("publicId", "")),
        "role": "",
        "event_time": "",
        "pay": "",
        "eligibility": "",
    }


def gig_row_sort_key(row: dict):
    if re.fullmatch(r"\d{2}:\d{2}", row["start_time_sort"]):
        hour, minute = (int(part) for part in row["start_time_sort"].split(":"))
        time_minutes = hour * 60 + minute
    else:
        time_minutes = 24 * 60 + 1
    return (
        row["date"],
        time_minutes,
        row["location_priority"],
        row["city"].casefold(),
        row["raw_listing"].casefold(),
    )


def normalize_month_events(
    raw_events: list[dict], config: MonthConfig, scraped_at: str
) -> list[dict]:
    month_events = []
    for event in raw_events:
        metadata_date = clean_text(event.get("metadataDate", ""))
        day_cell_date = clean_text(event.get("dayCellDate", ""))
        if metadata_date and day_cell_date and metadata_date != day_cell_date:
            print(
                "WARNING: FullCalendar start metadata and semantic day-cell date differ; "
                "using the day-cell data-date for this rendered month event. "
                f"metadata={metadata_date}, day-cell={day_cell_date}, "
                f"event={event.get('rawText', '')!r}"
            )
            event = dict(event)
            event["date"] = day_cell_date

        event_date = clean_text(event.get("date", ""))
        if not event_date.startswith(config.prefix + "-"):
            continue
        month_events.append(parse_calendar_listing(event, scraped_at, config))

    # FullCalendar instance/definition IDs are useful only for collapsing DOM
    # render clones in this run.  Only a genuine public event ID is stable
    # enough to distinguish otherwise identical gigs across daily runs.
    rows: list[dict] = []
    seen_public_ids: dict[tuple[str, str, str], tuple[str, str, str]] = {}
    seen_fallbacks_without_public_id: set[tuple[str, str, str]] = set()
    public_backed_fallbacks = {
        (
            row["date"],
            row["start_time_sort"],
            normalize_identity_text(row["raw_listing"]),
        )
        for row in month_events
        if clean_text(row.get("public_id", ""))
    }
    for row in month_events:
        fallback_key = (
            row["date"],
            row["start_time_sort"],
            normalize_identity_text(row["raw_listing"]),
        )
        public_id = clean_text(row.get("public_id", ""))
        if public_id:
            public_key = (public_id, row["date"], row["start_time_sort"])
            if public_key in seen_public_ids:
                if seen_public_ids[public_key] != fallback_key:
                    raise RuntimeError(
                        "A reliable calendar Event ID was reused by two conflicting "
                        f"gig listings in {config.label}: {public_key!r}."
                    )
                continue
            seen_public_ids[public_key] = fallback_key
            rows.append(row)
            continue

        # When no durable server-side ID exists, the user-requested composite
        # key is the authority.  A transient fcSeg instance ID must not make a
        # popover/month-grid clone look like a separate gig.
        if (
            fallback_key in public_backed_fallbacks
            or fallback_key in seen_fallbacks_without_public_id
        ):
            continue
        seen_fallbacks_without_public_id.add(fallback_key)
        rows.append(row)

    rows.sort(key=gig_row_sort_key)
    return rows


def excel_time_value(value: str):
    if not re.fullmatch(r"\d{2}:\d{2}", value):
        return None
    hour, minute = (int(part) for part in value.split(":"))
    return time(hour=hour, minute=minute)


def normalize_identity_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", clean_text(str(value or "")))
    return normalized.casefold()


def exact_fallback_key(row: dict) -> tuple[str, str, str]:
    return (
        clean_text(row.get("date", "")),
        clean_text(row.get("start_time_sort", "")),
        normalize_identity_text(row.get("raw_listing", "")),
    )


def structured_fallback_key(row: dict) -> tuple[str, ...]:
    return (
        clean_text(row.get("date", "")),
        clean_text(row.get("start_time_sort", "")),
        normalize_identity_text(row.get("state", "")),
        normalize_identity_text(row.get("store_account", "")),
        clean_text(row.get("store_number", "")),
        normalize_identity_text(row.get("city", "")),
    )


def canonical_excel_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(str(value or ""))
    if not text:
        return ""
    try:
        return date.fromisoformat(text[:10]).isoformat()
    except ValueError as exc:
        raise RuntimeError(f"Invalid historical Date value: {value!r}") from exc


def canonical_excel_time(value) -> str:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    text = clean_text(str(value or ""))
    if not text:
        return ""
    if re.fullmatch(r"\d{2}:\d{2}(?::\d{2})?", text):
        return text[:5]
    _, sortable = parse_dom_start_time(text)
    if sortable:
        return sortable
    _, sortable = parse_short_calendar_time(text)
    if sortable:
        return sortable
    raise RuntimeError(f"Invalid historical time value: {value!r}")


def canonical_history_timestamp(value) -> str:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return datetime.combine(value, time()).isoformat(timespec="seconds")
    text = clean_text(str(value or ""))
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).isoformat(timespec="seconds")
    except ValueError:
        for fmt in ("%Y-%m-%d %I:%M %p", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
            try:
                return datetime.strptime(text, fmt).isoformat(timespec="seconds")
            except ValueError:
                pass
    raise RuntimeError(f"Invalid historical timestamp value: {value!r}")


def load_prior_month_rows(config: MonthConfig) -> tuple[bool, list[dict]]:
    if not config.xlsx_path.exists():
        return False, []

    from openpyxl import load_workbook

    workbook = load_workbook(config.xlsx_path, data_only=False)
    try:
        sheet = workbook[config.sheet_name] if config.sheet_name in workbook.sheetnames else workbook.active
        headers = [clean_text(str(cell.value or "")) for cell in sheet[1]]
        if len(headers) != len(set(headers)):
            raise RuntimeError(f"Duplicate headers found in {config.xlsx_path.name}: {headers}")
        if headers not in (LEGACY_GIG_HEADERS, DAILY_GIG_HEADERS_V1, DAILY_GIG_HEADERS):
            raise RuntimeError(
                f"Unexpected schema in {config.xlsx_path.name}. Found headers: {headers}"
            )
        header_positions = {header: index + 1 for index, header in enumerate(headers)}
        prior_rows = []
        for row_number in range(2, sheet.max_row + 1):
            values = {
                header: sheet.cell(row=row_number, column=column).value
                for header, column in header_positions.items()
            }
            if not any(value not in (None, "") for value in values.values()):
                continue

            row_date = canonical_excel_date(values.get("Date"))
            if not row_date.startswith(config.prefix + "-"):
                raise RuntimeError(
                    f"Historical row {row_number} in {config.xlsx_path.name} is outside {config.label}: {row_date}"
                )
            start_sort = canonical_excel_time(values.get("Start Time Sort") or values.get("Start Time"))
            friendly_time = clean_text(str(values.get("Start Time") or ""))
            if isinstance(values.get("Start Time"), (datetime, time)):
                friendly_time = parse_dom_start_time(f"T{start_sort}:00")[0]

            scraped_at = canonical_history_timestamp(values.get("Scraped At"))
            is_legacy = headers == LEGACY_GIG_HEADERS
            status = "EXISTING" if is_legacy else clean_text(str(values.get("Status") or "")).upper()
            if status not in {"NEW", "EXISTING", "REMOVED"}:
                raise RuntimeError(
                    f"Invalid Status at row {row_number} in {config.xlsx_path.name}: {status!r}"
                )
            first_seen = scraped_at if is_legacy else canonical_history_timestamp(values.get("First Seen"))
            last_seen = scraped_at if is_legacy else canonical_history_timestamp(values.get("Last Seen"))
            if not is_legacy and (not first_seen or not last_seen):
                raise RuntimeError(
                    f"Missing First Seen/Last Seen at row {row_number} in {config.xlsx_path.name}."
                )

            city = clean_text(str(values.get("City") or ""))
            priority_value = values.get("Location Priority")
            priority = int(priority_value) if priority_value not in (None, "") else (
                {"bellevue": 1, "kirkland": 2, "seattle": 3}.get(city.casefold(), 9)
            )
            stable_event_id = clean_text(str(values.get("Event ID") or ""))
            prior_rows.append(
                {
                    "date": row_date,
                    "day": clean_text(str(values.get("Day") or date.fromisoformat(row_date).strftime("%A"))),
                    "start_time": friendly_time,
                    "start_time_sort": start_sort,
                    "state": clean_text(str(values.get("State") or "")),
                    "store_account": clean_text(str(values.get("Store / Account") or "")),
                    "store_number": clean_text(str(values.get("Store Number") or "")),
                    "city": city,
                    "preferred_area": clean_text(str(values.get("Preferred Area") or "NO")).upper(),
                    "location_priority": priority,
                    "raw_listing": clean_text(str(values.get("Raw Calendar Listing") or "")),
                    "calendar_color": clean_text(str(values.get("Calendar Color") or "")),
                    "event_url": clean_text(str(values.get("Event URL") or "")),
                    "status": status,
                    "first_seen": first_seen,
                    "last_seen": last_seen,
                    "scraped_at": scraped_at,
                    "event_id": "",
                    "public_id": stable_event_id,
                    "role": clean_text(str(values.get("Role") or "")),
                    "event_time": clean_text(str(values.get("Event Time") or "")),
                    "pay": clean_text(str(values.get("Pay") or "")),
                    "eligibility": clean_text(str(values.get("Eligibility") or "")),
                }
            )
    finally:
        workbook.close()

    return True, prior_rows


def match_rows_by_key(
    current_rows: list[dict],
    prior_rows: list[dict],
    matches: dict[int, int],
    used_prior: set[int],
    key_function,
) -> None:
    current_groups: dict[tuple, list[int]] = {}
    prior_groups: dict[tuple, list[int]] = {}
    for index, row in enumerate(current_rows):
        if index not in matches:
            current_groups.setdefault(key_function(row), []).append(index)
    for index, row in enumerate(prior_rows):
        if index not in used_prior:
            prior_groups.setdefault(key_function(row), []).append(index)

    for key in current_groups.keys() & prior_groups.keys():
        compatible_pairs = []
        for current_index in current_groups[key]:
            current_id = clean_text(current_rows[current_index].get("public_id", ""))
            for prior_index in prior_groups[key]:
                prior_id = clean_text(prior_rows[prior_index].get("public_id", ""))
                if current_id and prior_id and current_id != prior_id:
                    continue
                compatible_pairs.append((current_index, prior_index))
        current_candidates = {pair[0] for pair in compatible_pairs}
        prior_candidates = {pair[1] for pair in compatible_pairs}
        if len(current_candidates) == 1 and len(prior_candidates) == 1:
            current_index = next(iter(current_candidates))
            prior_index = next(iter(prior_candidates))
            matches[current_index] = prior_index
            used_prior.add(prior_index)
        elif compatible_pairs:
            raise RuntimeError(
                f"Ambiguous historical gig identity for key {key!r}; refusing to guess a match."
            )


def merge_daily_month_rows(
    current_rows: list[dict],
    prior_exists: bool,
    prior_rows: list[dict],
    run_at: str,
) -> tuple[list[dict], dict]:
    matches: dict[int, int] = {}
    used_prior: set[int] = set()
    current_id_counts = Counter(
        clean_text(row.get("public_id", "")) for row in current_rows if clean_text(row.get("public_id", ""))
    )
    prior_id_counts = Counter(
        clean_text(row.get("public_id", "")) for row in prior_rows if clean_text(row.get("public_id", ""))
    )
    prior_by_id = {
        clean_text(row.get("public_id", "")): index
        for index, row in enumerate(prior_rows)
        if clean_text(row.get("public_id", "")) and prior_id_counts[clean_text(row.get("public_id", ""))] == 1
    }

    for current_index, row in enumerate(current_rows):
        event_id = clean_text(row.get("public_id", ""))
        if event_id and current_id_counts[event_id] == 1 and event_id in prior_by_id:
            prior_index = prior_by_id[event_id]
            matches[current_index] = prior_index
            used_prior.add(prior_index)

    match_rows_by_key(current_rows, prior_rows, matches, used_prior, exact_fallback_key)
    match_rows_by_key(current_rows, prior_rows, matches, used_prior, structured_fallback_key)

    merged_rows = []
    for current_index, current in enumerate(current_rows):
        output = dict(current)
        prior_index = matches.get(current_index)
        if not prior_exists:
            output["status"] = "EXISTING"
            output["first_seen"] = run_at
        elif prior_index is None:
            output["status"] = "NEW"
            output["first_seen"] = run_at
        else:
            prior = prior_rows[prior_index]
            output["status"] = "NEW" if prior["status"] == "REMOVED" else "EXISTING"
            output["first_seen"] = prior["first_seen"]
            if not output.get("public_id"):
                output["public_id"] = prior.get("public_id", "")
            # Carry gig-detail data forward so already-fetched pay info survives
            # runs where the detail view was not re-opened.
            for field in DETAIL_FIELDS:
                if not clean_text(str(output.get(field) or "")):
                    output[field] = clean_text(str(prior.get(field) or ""))
        output["last_seen"] = run_at
        output["scraped_at"] = run_at
        merged_rows.append(output)

    for prior_index, prior in enumerate(prior_rows):
        if prior_index in used_prior:
            continue
        removed = dict(prior)
        removed["status"] = "REMOVED"
        removed["scraped_at"] = run_at
        merged_rows.append(removed)

    merged_rows.sort(key=gig_row_sort_key)
    summary = {
        "current": len(current_rows),
        "new": sum(row["status"] == "NEW" for row in merged_rows),
        "existing": sum(row["status"] == "EXISTING" for row in merged_rows),
        "removed": sum(row["status"] == "REMOVED" for row in merged_rows),
    }
    if summary["current"] != summary["new"] + summary["existing"]:
        raise RuntimeError(f"Daily merge summary invariant failed: {summary}")
    return merged_rows, summary


def validate_identity_uniqueness(rows: list[dict], context: str) -> None:
    seen_public_keys: set[tuple[str, str, str]] = set()
    public_fallbacks: set[tuple[str, str, str]] = set()
    idless_fallbacks: set[tuple[str, str, str]] = set()

    for row in rows:
        fallback_key = exact_fallback_key(row)
        public_id = clean_text(row.get("public_id", ""))
        if public_id:
            public_key = (
                public_id,
                clean_text(row.get("date", "")),
                clean_text(row.get("start_time_sort", "")),
            )
            if public_key in seen_public_keys:
                raise RuntimeError(
                    f"Duplicate reliable Event ID survived in {context}: {public_key!r}."
                )
            seen_public_keys.add(public_key)
            public_fallbacks.add(fallback_key)
        else:
            if fallback_key in idless_fallbacks:
                raise RuntimeError(
                    f"Duplicate composite gig key survived in {context}: {fallback_key!r}."
                )
            idless_fallbacks.add(fallback_key)

    overlap = public_fallbacks & idless_fallbacks
    if overlap:
        raise RuntimeError(
            f"An ID-less calendar rendering duplicates an ID-backed gig in {context}: "
            f"{next(iter(overlap))!r}."
        )


def validate_and_print_current_month(rows: list[dict], config: MonthConfig) -> None:
    validate_identity_uniqueness(rows, f"current {config.label} scrape")
    print(f"TOTAL {config.label.upper()} GIGS FOUND: {len(rows)}")
    previous_key = None
    for row in rows:
        if not row["date"].startswith(config.prefix + "-"):
            raise RuntimeError(f"Non-{config.label} row reached validation: {row}")
        current_key = gig_row_sort_key(row)
        if previous_key is not None and current_key < previous_key:
            raise RuntimeError(f"{config.label} current gigs are not sorted correctly.")
        previous_key = current_key
        store_display = " ".join(
            part for part in [row["store_account"], row["store_number"]] if part
        )
        print(f"{row['date']} | {row['start_time']} | {row['city']} | {store_display}")
    counts = Counter(row["date"] for row in rows)
    for event_date in sorted(counts):
        parsed = date.fromisoformat(event_date)
        print(f"{parsed.strftime('%B')} {parsed.day}: {counts[event_date]} gigs")


def validate_daily_history_rows(rows: list[dict], config: MonthConfig) -> None:
    validate_identity_uniqueness(rows, f"merged {config.label} history")
    previous_key = None
    preferred_priorities = {"bellevue": 1, "kirkland": 2, "seattle": 3}
    for row in rows:
        if not row["date"].startswith(config.prefix + "-"):
            raise RuntimeError(f"Historical row outside {config.label}: {row}")
        if row.get("status") not in {"NEW", "EXISTING", "REMOVED"}:
            raise RuntimeError(f"Invalid daily Status: {row}")
        if not row.get("first_seen") or not row.get("last_seen"):
            raise RuntimeError(f"Missing history timestamp: {row}")
        first_seen = history_timestamp_to_excel(row["first_seen"])
        last_seen = history_timestamp_to_excel(row["last_seen"])
        scraped_at = history_timestamp_to_excel(row.get("scraped_at", ""))
        if first_seen is None or last_seen is None or scraped_at is None:
            raise RuntimeError(f"Invalid history timestamp: {row}")
        if first_seen > last_seen:
            raise RuntimeError(f"First Seen is later than Last Seen: {row}")
        expected_priority = preferred_priorities.get(row["city"].casefold(), 9)
        expected_preferred = "YES" if expected_priority in {1, 2, 3} else "NO"
        if (
            row["location_priority"] != expected_priority
            or row["preferred_area"] != expected_preferred
        ):
            raise RuntimeError(f"Preferred-area classification is inconsistent: {row}")
        current_key = gig_row_sort_key(row)
        if previous_key is not None and current_key < previous_key:
            raise RuntimeError(f"{config.label} history is not sorted correctly.")
        previous_key = current_key


def daily_row_to_csv_values(row: dict) -> dict:
    return {
        "Date": row["date"],
        "Day": row["day"],
        "Start Time": row["start_time"],
        "Start Time Sort": row["start_time_sort"],
        "State": row["state"],
        "Store / Account": row["store_account"],
        "Store Number": row["store_number"],
        "City": row["city"],
        "Preferred Area": row["preferred_area"],
        "Location Priority": row["location_priority"],
        "Raw Calendar Listing": row["raw_listing"],
        "Calendar Color": row["calendar_color"],
        "Event URL": row["event_url"],
        "Status": row["status"],
        "First Seen": row["first_seen"],
        "Last Seen": row["last_seen"],
        "Scraped At": row["scraped_at"],
        "Event ID": row.get("public_id", ""),
        "Role": row.get("role", ""),
        "Event Time": row.get("event_time", ""),
        "Pay": row.get("pay", ""),
        "Eligibility": row.get("eligibility", ""),
    }


def history_timestamp_to_excel(value: str):
    if not value:
        return None
    return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)


def write_month_stage(
    rows: list[dict], config: MonthConfig, xlsx_stage: Path, csv_stage: Path
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo

    with csv_stage.open("w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=DAILY_GIG_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow(daily_row_to_csv_values(row))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = config.sheet_name
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.append(DAILY_GIG_HEADERS)

    for row in rows:
        start_value = excel_time_value(row["start_time_sort"])
        sheet.append(
            [
                date.fromisoformat(row["date"]),
                row["day"],
                start_value,
                start_value,
                row["state"],
                row["store_account"],
                row["store_number"],
                row["city"],
                row["preferred_area"],
                row["location_priority"],
                row["raw_listing"],
                row["calendar_color"],
                row["event_url"],
                row["status"],
                history_timestamp_to_excel(row["first_seen"]),
                history_timestamp_to_excel(row["last_seen"]),
                history_timestamp_to_excel(row["scraped_at"]),
                row.get("public_id", ""),
                row.get("role", ""),
                row.get("event_time", ""),
                row.get("pay", ""),
                row.get("eligibility", ""),
            ]
        )

    last_row = sheet.max_row
    if rows:
        table = Table(displayName=config.table_name, ref=f"A1:V{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        sheet.auto_filter.ref = f"A1:V{last_row}"
    else:
        sheet.auto_filter.ref = "A1:V1"

    dark_blue = "1F4E78"
    header_fill = PatternFill("solid", fgColor=dark_blue)
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    preferred_fill = PatternFill("solid", fgColor="E2F0D9")
    preferred_font = Font(color="375623", bold=True)
    new_fill = PatternFill("solid", fgColor="FFF2CC")
    removed_fill = PatternFill("solid", fgColor="F4CCCC")

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=dark_blue))
    sheet.row_dimensions[1].height = 32

    for row_number in range(2, last_row + 1):
        for column_number in range(1, len(DAILY_GIG_HEADERS) + 1):
            cell = sheet.cell(row=row_number, column=column_number)
            cell.alignment = Alignment(vertical="top", wrap_text=column_number in {11, 13})
            cell.border = Border(bottom=thin_gray)
        status = sheet.cell(row=row_number, column=14).value
        if status == "EXISTING" and sheet.cell(row=row_number, column=9).value == "YES":
            sheet.cell(row=row_number, column=9).fill = preferred_fill
            sheet.cell(row=row_number, column=9).font = preferred_font
        if status in {"NEW", "REMOVED"}:
            fill = new_fill if status == "NEW" else removed_fill
            for column_number in range(1, len(DAILY_GIG_HEADERS) + 1):
                sheet.cell(row=row_number, column=column_number).fill = fill
            status_cell = sheet.cell(row=row_number, column=14)
            status_cell.font = Font(
                color="7F6000" if status == "NEW" else "9C0006", bold=True
            )
        url_cell = sheet.cell(row=row_number, column=13)
        if isinstance(url_cell.value, str) and url_cell.value.startswith(("http://", "https://")):
            url_cell.hyperlink = url_cell.value
            url_cell.font = Font(color="0563C1", underline="single")

    for cell in sheet["A"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in sheet["C"][1:]:
        cell.number_format = "h:mm AM/PM"
    for cell in sheet["D"][1:]:
        cell.number_format = "hh:mm"
    for column in ("O", "P", "Q"):
        for cell in sheet[column][1:]:
            cell.number_format = "yyyy-mm-dd h:mm AM/PM"

    widths = {
        "A": 12, "B": 11, "C": 13, "D": 15, "E": 8, "F": 24,
        "G": 14, "H": 22, "I": 15, "J": 17, "K": 48, "L": 18,
        "M": 42, "N": 12, "O": 22, "P": 22, "Q": 22, "R": 28,
        "S": 20, "T": 22, "U": 16, "V": 34,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    workbook.save(xlsx_stage)


def verify_month_stage(
    rows: list[dict], config: MonthConfig, xlsx_stage: Path, csv_stage: Path
) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_stage, data_only=False)
    try:
        sheet = workbook[config.sheet_name]
        headers = [sheet.cell(1, column).value for column in range(1, len(DAILY_GIG_HEADERS) + 1)]
        if headers != DAILY_GIG_HEADERS:
            raise RuntimeError(f"{config.label} staged XLSX header verification failed.")
        if (
            sheet.max_row - 1 != len(rows)
            or sheet.max_column != len(DAILY_GIG_HEADERS)
            or str(sheet.freeze_panes) != "A2"
        ):
            raise RuntimeError(f"{config.label} staged XLSX structure verification failed.")
        expected_ref = f"A1:V{len(rows) + 1}" if rows else "A1:V1"
        if sheet.auto_filter.ref != expected_ref:
            raise RuntimeError(f"{config.label} staged autofilter verification failed.")
        if rows:
            if config.table_name not in sheet.tables or sheet.tables[config.table_name].ref != expected_ref:
                raise RuntimeError(f"{config.label} staged table verification failed.")
        for row_number, expected in enumerate(rows, start=2):
            status = sheet.cell(row=row_number, column=14).value
            if status != expected["status"]:
                raise RuntimeError(f"{config.label} Status verification failed at row {row_number}.")
            loaded_date = canonical_excel_date(sheet.cell(row=row_number, column=1).value)
            loaded_start = canonical_excel_time(sheet.cell(row=row_number, column=3).value)
            loaded_sort = canonical_excel_time(sheet.cell(row=row_number, column=4).value)
            if loaded_date != expected["date"]:
                raise RuntimeError(f"{config.label} Date value verification failed at row {row_number}.")
            if loaded_start != expected["start_time_sort"] or loaded_sort != expected["start_time_sort"]:
                raise RuntimeError(f"{config.label} time value verification failed at row {row_number}.")

            text_columns = {
                2: expected["day"],
                5: expected["state"],
                6: expected["store_account"],
                7: expected["store_number"],
                8: expected["city"],
                9: expected["preferred_area"],
                11: expected["raw_listing"],
                12: expected["calendar_color"],
                13: expected["event_url"],
                14: expected["status"],
                18: expected.get("public_id", ""),
                19: expected.get("role", ""),
                20: expected.get("event_time", ""),
                21: expected.get("pay", ""),
                22: expected.get("eligibility", ""),
            }
            for column_number, expected_value in text_columns.items():
                loaded_value = clean_text(str(sheet.cell(row=row_number, column=column_number).value or ""))
                if loaded_value != clean_text(str(expected_value or "")):
                    raise RuntimeError(
                        f"{config.label} value verification failed at row {row_number}, "
                        f"column {column_number}."
                    )
            if sheet.cell(row=row_number, column=10).value != expected["location_priority"]:
                raise RuntimeError(
                    f"{config.label} Location Priority verification failed at row {row_number}."
                )
            for column_number, expected_key in ((15, "first_seen"), (16, "last_seen"), (17, "scraped_at")):
                loaded_timestamp = canonical_history_timestamp(
                    sheet.cell(row=row_number, column=column_number).value
                )
                if loaded_timestamp != canonical_history_timestamp(expected[expected_key]):
                    raise RuntimeError(
                        f"{config.label} {expected_key} verification failed at row {row_number}."
                    )
            if sheet.cell(row=row_number, column=1).number_format != "yyyy-mm-dd":
                raise RuntimeError(f"{config.label} Date format verification failed.")
            if sheet.cell(row=row_number, column=3).number_format != "h:mm AM/PM" or \
                    sheet.cell(row=row_number, column=4).number_format != "hh:mm":
                raise RuntimeError(f"{config.label} time format verification failed.")
            fill = sheet.cell(row=row_number, column=1).fill.fgColor.rgb or ""
            if status == "NEW" and not fill.endswith("FFF2CC"):
                raise RuntimeError(f"{config.label} NEW-row highlight verification failed.")
            if status == "REMOVED" and not fill.endswith("F4CCCC"):
                raise RuntimeError(f"{config.label} REMOVED-row highlight verification failed.")
    finally:
        workbook.close()

    with csv_stage.open("r", newline="", encoding="utf-8-sig") as csvfile:
        reader = csv.DictReader(csvfile)
        csv_rows = list(reader)
        fieldnames = reader.fieldnames
    expected_csv = [
        {key: str(value) for key, value in daily_row_to_csv_values(row).items()}
        for row in rows
    ]
    if fieldnames != DAILY_GIG_HEADERS or csv_rows != expected_csv:
        raise RuntimeError(f"{config.label} staged CSV verification failed.")


def stage_month_outputs(rows_by_month: dict[str, list[dict]]) -> dict[str, dict]:
    staged = {}
    try:
        for config in MONTH_CONFIGS:
            token = uuid4().hex
            xlsx_stage = ROOT / f".{config.xlsx_path.stem}.{token}.tmp.xlsx"
            csv_stage = ROOT / f".{config.csv_path.stem}.{token}.tmp.csv"
            rows = rows_by_month[config.prefix]
            staged[config.prefix] = {
                "config": config,
                "xlsx_stage": xlsx_stage,
                "csv_stage": csv_stage,
            }
            write_month_stage(rows, config, xlsx_stage, csv_stage)
            verify_month_stage(rows, config, xlsx_stage, csv_stage)
        return staged
    except Exception:
        for item in staged.values():
            item["xlsx_stage"].unlink(missing_ok=True)
            item["csv_stage"].unlink(missing_ok=True)
        raise


def timestamped_backup_path(config: MonthConfig, run_at: str) -> Path:
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = history_timestamp_to_excel(run_at).strftime("%Y-%m-%d_%H%M%S")
    candidate = BACKUP_DIR / f"{config.xlsx_path.stem}_{stamp}.xlsx"
    suffix = 1
    while candidate.exists():
        candidate = BACKUP_DIR / f"{config.xlsx_path.stem}_{stamp}_{suffix}.xlsx"
        suffix += 1
    return candidate


def commit_staged_outputs(staged: dict[str, dict], run_at: str) -> None:
    rollback_paths: dict[Path, Path] = {}
    originally_missing: set[Path] = set()
    targets = []
    try:
        for config in MONTH_CONFIGS:
            item = staged[config.prefix]
            for target in (config.xlsx_path, config.csv_path):
                targets.append(target)
                if target.exists():
                    rollback = ROOT / f".{target.name}.{uuid4().hex}.rollback"
                    shutil.copy2(target, rollback)
                    rollback_paths[target] = rollback
                else:
                    originally_missing.add(target)
            if config.xlsx_path.exists():
                backup = timestamped_backup_path(config, run_at)
                shutil.copy2(config.xlsx_path, backup)
                print(f"Created workbook backup: {backup}")
            os.replace(item["xlsx_stage"], config.xlsx_path)
            os.replace(item["csv_stage"], config.csv_path)
    except Exception:
        for target in targets:
            if target in rollback_paths:
                os.replace(rollback_paths[target], target)
            elif target in originally_missing:
                target.unlink(missing_ok=True)
        raise
    finally:
        for rollback in rollback_paths.values():
            rollback.unlink(missing_ok=True)
        for item in staged.values():
            item["xlsx_stage"].unlink(missing_ok=True)
            item["csv_stage"].unlink(missing_ok=True)

    for config in MONTH_CONFIGS:
        print(f"Updated Excel spreadsheet: {config.xlsx_path}")
        print(f"Updated CSV export: {config.csv_path}")


def print_daily_sync_summaries(results: dict[str, dict]) -> None:
    print("\nDAILY SYNC SUMMARY")
    for config in MONTH_CONFIGS:
        result = results[config.prefix]
        summary = result["summary"]
        print(f"\n{config.label.upper()}")
        print(f"Current gigs on site: {summary['current']}")
        print(f"New gigs: {summary['new']}")
        print(f"Existing gigs: {summary['existing']}")
        print(f"Removed gigs: {summary['removed']}")

    for status in ("NEW", "REMOVED"):
        print(f"\n{status} GIGS")
        printed = False
        for config in MONTH_CONFIGS:
            for row in results[config.prefix]["rows"]:
                if row["status"] != status:
                    continue
                printed = True
                store_display = " ".join(
                    part for part in [row["store_account"], row["store_number"]] if part
                )
                print(
                    f"{status} | {row['date']} | {row['start_time']} | "
                    f"{row['city']} | {store_display}"
                )
        if not printed:
            print("(none)")


def launch_browser(p: Playwright):
    launch_args = ["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"]
    browser = p.chromium.launch(headless=False, args=launch_args)
    print("Using default Chromium.")
    return browser


def seed_details_from_prior(rows: list[dict], prior_rows: list[dict]) -> None:
    """Copy previously fetched gig-detail fields onto current rows by Event ID."""
    prior_by_id = {
        clean_text(prior.get("public_id", "")): prior
        for prior in prior_rows
        if clean_text(prior.get("public_id", "")) and clean_text(prior.get("pay", ""))
    }
    for row in rows:
        prior = prior_by_id.get(clean_text(row.get("public_id", "")))
        if prior and not clean_text(row.get("pay", "")):
            for field in DETAIL_FIELDS:
                row[field] = clean_text(str(prior.get(field) or ""))


def mark_gig_event_for_detail_click(page, row: dict) -> bool:
    """Tag the rendered calendar element for this gig so it can be clicked.

    Returns True when a unique clickable element was marked. Events hidden
    behind a "+N more" overflow are exposed by opening that day's popover.
    """

    def try_mark() -> bool:
        return bool(
            page.evaluate(
                """
                target => {
                  const clean = value => String(value || '').replace(/\\s+/g, ' ').trim();
                  document.querySelectorAll('[data-us-event-detail-target]').forEach(el =>
                    el.removeAttribute('data-us-event-detail-target'));
                  const matches = [...document.querySelectorAll('*')].filter(el => {
                    const seg = el.fcSeg;
                    if (!seg || seg.isStart === false) return false;
                    const def = seg.eventRange?.def || {};
                    const props = def.extendedProps || {};
                    const ids = [
                      def.publicId, props.eventId, props.event_id, props.gigId,
                      props.gig_id, props.shiftId, props.shift_id,
                    ].map(clean).filter(Boolean);
                    if (target.publicId && ids.includes(target.publicId)) return true;
                    if (target.publicId) return false;
                    // Fallback identity for events without a durable server ID.
                    const cellDate = el.closest('[data-date]')?.getAttribute('data-date') || '';
                    return cellDate === target.date &&
                      clean(el.textContent).toLowerCase() === target.rawText.toLowerCase();
                  });
                  const visible = matches.filter(el => el.getClientRects().length > 0);
                  const chosen = visible[0] || matches[0];
                  if (chosen) chosen.setAttribute('data-us-event-detail-target', 'true');
                  return Boolean(chosen);
                }
                """,
                {
                    "publicId": clean_text(row.get("public_id", "")),
                    "date": row["date"],
                    "rawText": clean_text(row.get("raw_listing", "")),
                },
            )
        )

    if try_mark():
        return True

    # The event may be hidden behind that day's "+N more" overflow control.
    for link in discover_overflow_links(page):
        if link["date"] != row["date"]:
            continue
        if mark_overflow_link_for_click(page, link) != 1:
            continue
        page.locator('[data-us-event-overflow-target="true"]').click(timeout=10000)
        page.wait_for_timeout(500)
        if try_mark():
            return True
        close_calendar_overflow_popover(page)
    return False


def extract_gig_detail_sections(page) -> list[dict]:
    """Extract role/time/pay/eligibility cards from an open gig detail view."""
    return page.evaluate(
        """
        () => {
          const clean = value => String(value || '').replace(/\\s+/g, ' ').trim();
          const isVisible = el => {
            if (!el.getClientRects().length) return false;
            const style = getComputedStyle(el);
            return style.display !== 'none' && style.visibility !== 'hidden';
          };
          // Material icon ligature names (e.g. "access_time", "attach_money")
          // leak into textContent; they are never real role/label text.
          const isIconText = (el, text) =>
            /icon/i.test(el.getAttribute('class') || '') || /^[a-z][a-z0-9_]*$/.test(text);
          const timeRe = /\\b\\d{1,2}(?::\\d{2})?\\s*[ap]\\.?m\\.?\\s*[-\\u2013\\u2014]\\s*\\d{1,2}(?::\\d{2})?\\s*[ap]\\.?m\\.?/i;
          const payRe = /\\$\\s*\\d[\\d,]*(?:\\.\\d{1,2})?\\s*(?:\\([^)]{0,40}\\))?(?:\\s*(?:\\/|per)\\s*(?:hr|hour))?/i;
          const buttonRe = /^(apply|applied|not available|unavailable)$/i;
          // Only the active detail tab: hidden panels from previously opened
          // gigs stay in the SPA DOM and must not contribute sections.
          const buttons = [...document.querySelectorAll('button, [role="button"], a, input[type="button"], input[type="submit"]')]
            .filter(el => isVisible(el) && buttonRe.test(clean(el.textContent || el.value)));
          const sections = [];
          const seen = new Set();
          for (const button of buttons) {
            let node = button.parentElement;
            let container = null;
            for (let depth = 0; depth < 10 && node; depth++) {
              const text = node.textContent || '';
              if (payRe.test(text) || timeRe.test(text)) { container = node; break; }
              node = node.parentElement;
            }
            if (!container || seen.has(container)) continue;
            seen.add(container);
            const text = clean(container.textContent);
            const pay = clean((text.match(payRe) || [''])[0]);
            const time = clean((text.match(timeRe) || [''])[0]);
            const eligibility = clean(
              (text.match(/you are[^.$]{0,120}?eligib[^.$]{0,60}/i) || [''])[0]
            );
            // Role: prefer a short leaf label inside the card that is not
            // time/pay/button text, then fall back to preceding sibling headers
            // (e.g. the "Brand Ambassador" banner above the card).
            let role = '';
            const leafEntries = [...container.querySelectorAll('*')]
              .filter(el => el.children.length === 0 && isVisible(el))
              .map(el => [el, clean(el.textContent)])
              .filter(([, t]) => Boolean(t));
            role = (leafEntries.find(([el, t]) =>
              t.length < 60 && !payRe.test(t) && !timeRe.test(t) &&
              !buttonRe.test(t) && !/eligib/i.test(t) && /[a-z]/i.test(t) &&
              !/^\\d/.test(t) && !isIconText(el, t)) || [null, ''])[1];
            let probe = container;
            for (let depth = 0; depth < 6 && probe && !role; depth++) {
              let sibling = probe.previousElementSibling;
              while (sibling && !role) {
                const t = clean(sibling.textContent);
                if (t && t.length < 60 && isVisible(sibling) && !payRe.test(t) &&
                    !timeRe.test(t) && !isIconText(sibling, t)) role = t;
                sibling = sibling.previousElementSibling;
              }
              probe = probe.parentElement;
            }
            sections.push({
              role: clean(role),
              time,
              pay,
              eligibility,
              buttonText: clean(button.textContent || button.value),
            });
          }
          if (!sections.length) {
            const body = clean(document.body.textContent);
            const pay = clean((body.match(payRe) || [''])[0]);
            const time = clean((body.match(timeRe) || [''])[0]);
            if (pay) sections.push({ role: '', time, pay, eligibility: '', buttonText: '' });
          }
          return sections;
        }
        """
    )


def return_to_month_calendar(page, config: MonthConfig) -> None:
    """Get back to the correct month's calendar after viewing a gig detail."""
    tab = page.get_by_role("tab", name=re.compile(r"^\s*calendar\s*$", re.I))
    if tab.count() == 0:
        tab = page.get_by_text(re.compile(r"^\s*Calendar\s*$"))
    tab.first.click(timeout=10000)
    page.wait_for_timeout(700)
    try:
        wait_for_month_calendar(page, config)
    except Exception:
        navigate_to_calendar_month(page, config)
        wait_for_month_calendar(page, config)


def fetch_missing_event_details(page, config: MonthConfig, rows: list[dict]) -> None:
    """Open each gig lacking pay info and record role/time/pay/eligibility.

    Controlled by env vars:
    - US_EVENT_SKIP_DETAILS=1 skips detail fetching entirely.
    - US_EVENT_DETAIL_LIMIT=N caps how many gigs are opened per month.
    Failures are per-gig and non-fatal: the row keeps blank detail fields and
    is retried on the next run.
    """
    if os.getenv("US_EVENT_SKIP_DETAILS", "").strip() == "1":
        print(f"{config.label}: gig detail fetching skipped (US_EVENT_SKIP_DETAILS=1).")
        return
    limit_env = os.getenv("US_EVENT_DETAIL_LIMIT", "").strip()
    limit = int(limit_env) if limit_env else None

    pending = [row for row in rows if not clean_text(row.get("pay", ""))]
    if limit is not None:
        pending = pending[:limit]
    if not pending:
        print(f"{config.label}: all gigs already have pay details.")
        return
    print(f"{config.label}: fetching details for {len(pending)} gig(s)...")

    for index, row in enumerate(pending, start=1):
        gig_label = f"{row['date']} {row.get('raw_listing', '')}"
        try:
            if not mark_gig_event_for_detail_click(page, row):
                print(f"  [{index}/{len(pending)}] SKIP (element not found): {gig_label}")
                continue
            page.locator('[data-us-event-detail-target="true"]').click(timeout=10000)
            page.wait_for_function(
                """
                () => {
                  const clean = v => String(v || '').replace(/\\s+/g, ' ').trim();
                  const isVisible = el => {
                    if (!el.getClientRects().length) return false;
                    const style = getComputedStyle(el);
                    return style.display !== 'none' && style.visibility !== 'hidden';
                  };
                  return [...document.querySelectorAll('button, [role="button"], a, input[type="button"], input[type="submit"]')]
                    .some(el => isVisible(el) &&
                      /^(apply|applied|not available|unavailable)$/i.test(clean(el.textContent || el.value)));
                }
                """,
                timeout=20000,
            )
            page.wait_for_timeout(600)
            sections = extract_gig_detail_sections(page)
            if sections:
                row["role"] = " | ".join(
                    dict.fromkeys(s["role"] for s in sections if s["role"])
                )
                row["event_time"] = " | ".join(
                    dict.fromkeys(s["time"] for s in sections if s["time"])
                )
                row["pay"] = " | ".join(
                    dict.fromkeys(s["pay"] for s in sections if s["pay"])
                )
                row["eligibility"] = " | ".join(
                    dict.fromkeys(s["eligibility"] for s in sections if s["eligibility"])
                )
                print(
                    f"  [{index}/{len(pending)}] {gig_label} -> "
                    f"role={row['role'] or '?'} time={row['event_time'] or '?'} pay={row['pay'] or '?'}"
                )
            else:
                print(f"  [{index}/{len(pending)}] WARN no role sections found: {gig_label}")
                save_debug_screenshot(page, "detail_no_sections")
        except Exception as exc:
            print(f"  [{index}/{len(pending)}] WARN detail fetch failed for {gig_label}: {exc}")
            save_debug_screenshot(page, "detail_fetch_failed")
        finally:
            try:
                return_to_month_calendar(page, config)
            except Exception as exc:
                raise RuntimeError(
                    f"Could not return to the {config.label} calendar after a gig "
                    f"detail view: {exc}"
                ) from exc

    fetched = sum(1 for row in pending if clean_text(row.get("pay", "")))
    print(f"{config.label}: detail fetch complete ({fetched}/{len(pending)} gigs got pay info).")


def main() -> None:
    with sync_playwright() as p:  # type: ignore[call-overload]
        browser = launch_browser(p)
        storage_state = str(AUTH_STATE_PATH) if AUTH_STATE_PATH.exists() else None
        context = browser.new_context(viewport={"width": 1440, "height": 1100}, storage_state=storage_state)
        page = context.new_page()

        try:
            login(page)
            page.wait_for_url(HOME_URL, timeout=20000)
            print("Reached /home successfully.")
            print("Pausing 8 seconds for the dashboard to settle...")
            page.wait_for_timeout(8000)

            wait_for_dashboard(page)
            context.storage_state(path=str(AUTH_STATE_PATH))
            print(f"Saved browser auth state to: {AUTH_STATE_PATH}")

            click_top_header_calendar(page)
            run_at = datetime.now().astimezone().replace(tzinfo=None).isoformat(timespec="seconds")
            current_rows_by_month: dict[str, list[dict]] = {}
            prior_by_month: dict[str, tuple[bool, list[dict]]] = {}

            # Scrape and validate every configured month before modifying any
            # workbook. This keeps the four-month refresh transactional.
            for config in MONTH_CONFIGS:
                try:
                    navigate_to_calendar_month(page, config)
                    wait_for_month_calendar(page, config)
                    inspect_calendar_dom(page, config)
                    raw_events = collect_all_month_dom_events(page, config)
                    rows = normalize_month_events(raw_events, config, run_at)
                    prior_exists, prior_rows = load_prior_month_rows(config)
                    prior_by_month[config.prefix] = (prior_exists, prior_rows)
                    seed_details_from_prior(rows, prior_rows)
                    fetch_missing_event_details(page, config, rows)
                    validate_and_print_current_month(rows, config)
                    current_rows_by_month[config.prefix] = rows
                except Exception as exc:
                    raise RuntimeError(
                        f"Failed processing {config.label}: {exc}"
                    ) from exc

            results: dict[str, dict] = {}
            rows_to_export: dict[str, list[dict]] = {}
            for config in MONTH_CONFIGS:
                prior_exists, prior_rows = prior_by_month[config.prefix]
                merged_rows, summary = merge_daily_month_rows(
                    current_rows_by_month[config.prefix],
                    prior_exists,
                    prior_rows,
                    run_at,
                )
                validate_daily_history_rows(merged_rows, config)
                rows_to_export[config.prefix] = merged_rows
                results[config.prefix] = {"rows": merged_rows, "summary": summary}

            staged = stage_month_outputs(rows_to_export)
            commit_staged_outputs(staged, run_at)
            print_daily_sync_summaries(results)
            return

        except Exception as exc:
            print_error_and_exit(str(exc), page)
        finally:
            browser.close()


if __name__ == "__main__":
    main()
